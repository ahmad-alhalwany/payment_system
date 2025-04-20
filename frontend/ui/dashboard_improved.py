import requests
from PyQt6.QtWidgets import (
    QMainWindow, QTabWidget, QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QPushButton, QTableWidget, QTableWidgetItem, QHeaderView, QMessageBox, QDialog,
    QLineEdit, QFormLayout, QComboBox, QGroupBox, QGridLayout, QDateEdit,
    QDoubleSpinBox, QRadioButton, QButtonGroup, QFileDialog, QTextEdit,
    QStatusBar, QAbstractSpinBox
)
from PyQt6.QtGui import QFont, QColor, QAction
from PyQt6.QtCore import Qt, QDate
from ui.money_transfer_improved import MoneyTransferApp
from ui.branch_management_improved import AddBranchDialog
from ui.user_management_improved import AddEmployeeDialog
from ui.user_search import UserSearchDialog
from ui.arabic_amount import number_to_arabic_words


import datetime
import os
import json

class ModernGroupBox(QGroupBox):
    """Custom styled group box."""
    
    def __init__(self, title, color="#3498db"):
        super().__init__(title)
        self.setStyleSheet(f"""
            QGroupBox {{
                font-weight: bold;
                border: 1px solid {color};
                border-radius: 5px;
                margin-top: 1em;
                padding-top: 10px;
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
                color: {color};
            }}
        """)

class ModernButton(QPushButton):
    """Custom styled button."""
    
    def __init__(self, text, color="#3498db"):
        super().__init__(text)
        self.setStyleSheet(f"""
            QPushButton {{
                background-color: {color};
                color: white;
                border-radius: 5px;
                padding: 8px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background-color: {self.lighten_color(color)};
            }}
            QPushButton:pressed {{
                background-color: {self.darken_color(color)};
            }}
        """)
    
    def lighten_color(self, color):
        """Lighten a hex color."""
        # Simple implementation - not perfect but works for our needs
        if color.startswith('#'):
            color = color[1:]
        r = min(255, int(color[0:2], 16) + 20)
        g = min(255, int(color[2:4], 16) + 20)
        b = min(255, int(color[4:6], 16) + 20)
        return f"#{r:02x}{g:02x}{b:02x}"
    
    def darken_color(self, color):
        """Darken a hex color."""
        if color.startswith('#'):
            color = color[1:]
        r = max(0, int(color[0:2], 16) - 20)
        g = max(0, int(color[2:4], 16) - 20)
        b = max(0, int(color[4:6], 16) - 20)
        return f"#{r:02x}{g:02x}{b:02x}"

class DirectorDashboard(QMainWindow):
    """Dashboard for the director role."""
    
    def __init__(self, token=None):
        super().__init__()
        self.token = token
        self.api_url = os.environ["API_URL"]
        self.current_page = 1
        self.total_pages = 1
        self.per_page = 7
        self.current_page_transactions = 1
        self.transactions_per_page = 15
        self.total_pages_transactions = 1
        
        self.setWindowTitle("لوحة تحكم المدير")
        self.setGeometry(100, 100, 1200, 800)
        
        # Create menu bar
        self.create_menu_bar()
        
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f5f5f5;
                font-family: Arial;
            }
            QLabel {
                color: #333;
            }
            QTabWidget::pane {
                border: 1px solid #ccc;
                border-radius: 8px;
                background-color: rgba(255, 255, 255, 0.7);
            }
            QTabBar::tab {
                background-color: #ddd;
                padding: 10px 15px;
                margin-right: 2px;
                border-top-left-radius: 5px;
                border-top-right-radius: 5px;
                font-weight: bold;
            }
            QTabBar::tab:selected {
                background-color: #2c3e50;
                color: white;
            }
        """)
        
        # Central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # Main layout
        main_layout = QVBoxLayout(central_widget)
        
        # Header
        header_layout = QHBoxLayout()
        
        # Logo/Title
        logo_label = QLabel("نظام التحويلات المالية")
        logo_label.setFont(QFont("Arial", 20, QFont.Weight.Bold))
        logo_label.setStyleSheet("color: #2c3e50;")
        header_layout.addWidget(logo_label)
        
        # Spacer
        header_layout.addStretch()
        
        # User info
        user_info = QLabel("مدير النظام")
        user_info.setFont(QFont("Arial", 12))
        user_info.setStyleSheet("color: #7f8c8d;")
        header_layout.addWidget(user_info)
        
        main_layout.addLayout(header_layout)
        
        # Tab widget
        self.tabs = QTabWidget()
        
        # Dashboard tab
        self.dashboard_tab = QWidget()
        self.setup_dashboard_tab()
        self.tabs.addTab(self.dashboard_tab, "لوحة المعلومات")
        
        # Branches tab
        self.branches_tab = QWidget()
        self.setup_branches_tab()
        self.tabs.addTab(self.branches_tab, "الفروع")
        
        # Employees tab
        self.employees_tab = QWidget()
        self.setup_employees_tab()
        self.tabs.addTab(self.employees_tab, "الموظفين")
        
        # Transactions tab
        self.transactions_tab = QWidget()
        self.setup_transactions_tab()
        self.tabs.addTab(self.transactions_tab, "التحويلات")
        
        # Admin Money Transfer tab
        self.admin_transfer_tab = QWidget()
        self.setup_admin_transfer_tab()
        self.tabs.addTab(self.admin_transfer_tab, "تحويل الأموال")
        
        # Reports tab
        self.reports_tab = QWidget()
        self.setup_reports_tab()
        self.tabs.addTab(self.reports_tab, "التقارير")
        
        # Settings tab
        self.settings_tab = QWidget()
        self.setup_settings_tab()
        self.tabs.addTab(self.settings_tab, "الإعدادات")
        
        main_layout.addWidget(self.tabs)
        
        # Status bar
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage("جاهز")
        
        # Load initial data
        self.load_dashboard_data()
        self.load_branches()
        self.load_employees()
        self.load_transactions()
        
        # Load branch filters for all tabs at once
        self.load_branches_for_filter()
        
        # Initialize branch_id_to_name mapping if not already done
        if not hasattr(self, 'branch_id_to_name'):
            self.branch_id_to_name = {}
    
    def create_menu_bar(self):
        """Create menu bar with logout and close options."""
        # Create menu bar
        menu_bar = self.menuBar()
        
        # Create user menu
        user_menu = menu_bar.addMenu("المستخدم")
        
        # Add logout action
        logout_action = QAction("تسجيل الخروج", self)
        logout_action.triggered.connect(self.logout)
        user_menu.addAction(logout_action)
        
        # Add separator
        user_menu.addSeparator()
        
        # Add close action
        close_action = QAction("إغلاق البرنامج", self)
        close_action.triggered.connect(self.close)
        user_menu.addAction(close_action)    
        
    def logout(self):
        """Logout and return to login screen."""
        reply = QMessageBox.question(
            self, 
            "تسجيل الخروج", 
            "هل أنت متأكد من رغبتك في تسجيل الخروج؟",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, 
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            self.close()
            # Signal to main application to show login window
            if self.parent():
                self.parent().show_login()
        
    def setup_admin_transfer_tab(self):
        """Set up the admin money transfer tab with unrestricted capabilities."""
        layout = QVBoxLayout()
        
        # Create the money transfer widget for System Manager with unrestricted capabilities
        self.admin_transfer_widget = MoneyTransferApp(
            user_token=self.token,
            branch_id=0,  # Main Branch with branch number 0
            user_id=0,    # System Manager ID is 0
            user_role="director",  # Special role for unrestricted access
            username="system_manager",
            full_name="System Manager",  # Set employee name to "System Manager"
        )
        
        layout.addWidget(self.admin_transfer_widget)
        self.admin_transfer_tab.setLayout(layout)    
    
    def setup_dashboard_tab(self):
        """Set up the dashboard tab with statistics and charts."""
        layout = QVBoxLayout()
        
        # Title
        title = QLabel("لوحة المعلومات")
        title.setFont(QFont("Arial", 18, QFont.Weight.Bold))
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setStyleSheet("color: #2c3e50; margin-bottom: 20px;")
        layout.addWidget(title)
        
        # Statistics cards
        stats_layout = QHBoxLayout()
        
        # Branches card
        self.branches_card = ModernGroupBox("الفروع", "#3498db")
        branches_layout = QVBoxLayout()
        self.branches_count = QLabel("0")
        self.branches_count.setFont(QFont("Arial", 24, QFont.Weight.Bold))
        self.branches_count.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.branches_count.setStyleSheet("color: #3498db;")
        branches_layout.addWidget(self.branches_count)
        branches_label = QLabel("إجمالي الفروع")
        branches_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        branches_layout.addWidget(branches_label)
        self.branches_card.setLayout(branches_layout)
        stats_layout.addWidget(self.branches_card)
        
        # Employees card
        self.employees_card = ModernGroupBox("الموظفين", "#2ecc71")
        employees_layout = QVBoxLayout()
        self.employees_count = QLabel("0")
        self.employees_count.setFont(QFont("Arial", 24, QFont.Weight.Bold))
        self.employees_count.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.employees_count.setStyleSheet("color: #2ecc71;")
        employees_layout.addWidget(self.employees_count)
        employees_label = QLabel("إجمالي الموظفين")
        employees_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        employees_layout.addWidget(employees_label)
        self.employees_card.setLayout(employees_layout)
        stats_layout.addWidget(self.employees_card)
        
        # Transactions card
        self.transactions_card = ModernGroupBox("التحويلات", "#e74c3c")
        transactions_layout = QVBoxLayout()
        self.transactions_count = QLabel("0")
        self.transactions_count.setFont(QFont("Arial", 24, QFont.Weight.Bold))
        self.transactions_count.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.transactions_count.setStyleSheet("color: #e74c3c;")
        transactions_layout.addWidget(self.transactions_count)
        transactions_label = QLabel("إجمالي التحويلات")
        transactions_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        transactions_layout.addWidget(transactions_label)
        self.transactions_card.setLayout(transactions_layout)
        stats_layout.addWidget(self.transactions_card)
        
        # Amount card
        self.amount_card = ModernGroupBox("المبالغ", "#f39c12")
        amount_layout = QVBoxLayout()
        self.amount_total = QLabel("0")
        self.amount_total.setFont(QFont("Arial", 24, QFont.Weight.Bold))
        self.amount_total.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.amount_total.setStyleSheet("color: #f39c12;")
        amount_layout.addWidget(self.amount_total)
        amount_label = QLabel("إجمالي المبالغ")
        amount_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        amount_layout.addWidget(amount_label)
        self.amount_card.setLayout(amount_layout)
        stats_layout.addWidget(self.amount_card)
        
        layout.addLayout(stats_layout)
        
        # Charts section - replaced with placeholder since matplotlib is not available
        charts_layout = QHBoxLayout()
        
        # Statistics by Branch
        stats_group = ModernGroupBox("إحصائيات الفروع", "#3498db")
        stats_layout = QHBoxLayout()
        
        # Transfers by Branch
        transfers_group = QGroupBox("التحويلات حسب الفرع")
        transfers_layout = QVBoxLayout()
        
        self.transfers_bars = QLabel()
        self.transfers_bars.setTextFormat(Qt.TextFormat.RichText)
        self.transfers_bars.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.transfers_bars.setStyleSheet("font-family: Arial; color: #333;")
        transfers_layout.addWidget(self.transfers_bars)
        
        transfers_group.setLayout(transfers_layout)
        stats_layout.addWidget(transfers_group)
        
        # Amounts by Branch
        amounts_group = QGroupBox("المبالغ حسب الفرع")
        amounts_layout = QVBoxLayout()
        
        self.amounts_bars = QLabel()
        self.amounts_bars.setTextFormat(Qt.TextFormat.RichText)
        self.amounts_bars.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.amounts_bars.setStyleSheet("font-family: Arial; color: #333;")
        amounts_layout.addWidget(self.amounts_bars)
        
        amounts_group.setLayout(amounts_layout)
        stats_layout.addWidget(amounts_group)

        stats_group.setLayout(stats_layout)
        layout.addWidget(stats_group)
        
        layout.addLayout(charts_layout)
        
        # Recent transactions
        recent_group = ModernGroupBox("أحدث التحويلات", "#3498db")
        recent_layout = QVBoxLayout()
        
        self.recent_transactions_table = QTableWidget()
        self.recent_transactions_table.setColumnCount(11)  # Increased to 11 columns
        self.recent_transactions_table.setHorizontalHeaderLabels([
            "النوع", "رقم التحويل", 
            "المرسل",  "المستلم", "المبلغ", "التاريخ", "الحالة", "الفرع الصادر",
            "اتجاه الصادر","الفرع الوارد", "اتجاه الوارد", "اسم الموظف"
        ])
        self.recent_transactions_table.horizontalHeader().setStretchLastSection(True)
        self.recent_transactions_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.recent_transactions_table.setStyleSheet("""
            QTableWidget {
                border: none;
                background-color: white;
                gridline-color: #ddd;
            }
            QHeaderView::section {
                background-color: #2c3e50;
                color: white;
                padding: 8px;
                border: 1px solid #1a2530;
                font-weight: bold;
            }
            QTableWidget::item {
                padding: 5px;
            }
            QTableWidget::item:selected {
                background-color: #3498db;
                color: white;
            }
            QTableWidget::item[type="outgoing"] {
                background-color: #e8f5e9;
            }
            QTableWidget::item[type="incoming"] {
                background-color: #ffebee;
            }
        """)
        
        # Pagination controls
        pagination_layout = QHBoxLayout()
        self.prev_button = ModernButton("السابق", color="#3498db")
        self.prev_button.clicked.connect(self.prev_page)
        pagination_layout.addWidget(self.prev_button)
        
        self.page_label = QLabel("الصفحة: 1")
        pagination_layout.addWidget(self.page_label)
        
        self.next_button = ModernButton("التالي", color="#3498db")
        self.next_button.clicked.connect(self.next_page)
        pagination_layout.addWidget(self.next_button)
        
        recent_layout.addWidget(self.recent_transactions_table)
        recent_layout.addLayout(pagination_layout)
        
        # View all button
        view_all_button = ModernButton("عرض جميع التحويلات", color="#3498db")
        view_all_button.clicked.connect(lambda: self.tabs.setCurrentIndex(3))  # Switch to transactions tab
        recent_layout.addWidget(view_all_button)
        
        recent_group.setLayout(recent_layout)
        layout.addWidget(recent_group)
        
        # Load branch statistics
        self.load_branch_stats()
        
        self.dashboard_tab.setLayout(layout)
        self.load_recent_transactions()
        
    def create_transaction_type_item(self, transaction):
        """Enhanced direction indicators"""
        transfer_type = ""
        color = QColor()
        
        has_outgoing = transaction.get("branch_id") is not None
        has_incoming = transaction.get("destination_branch_id") is not None
        
        if has_outgoing and has_incoming:
            if transaction["branch_id"] == transaction["destination_branch_id"]:
                transfer_type = "↔ داخلي"
                color = QColor(150, 150, 0)  # Yellow
            else:
                transfer_type = "↔ بين فروع"
                color = QColor(0, 0, 150)  # Blue
        elif has_outgoing:
            transfer_type = "↑ صادر"
            color = QColor(0, 150, 0)  # Green
        elif has_incoming:
            transfer_type = "↓ وارد"
            color = QColor(150, 0, 0)  # Red
        else:
            transfer_type = "↔ نظامي"
            color = QColor(100, 100, 100)  # Gray
        
        item = QTableWidgetItem(transfer_type)
        item.setForeground(color)
        item.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        return item
    
    def setup_branches_tab(self):
        """Set up the branches tab."""
        layout = QVBoxLayout()
        
        # Title
        title = QLabel("إدارة الفروع")
        title.setFont(QFont("Arial", 18, QFont.Weight.Bold))
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setStyleSheet("color: #2c3e50; margin-bottom: 20px;")
        layout.addWidget(title)
        
        # Branches table
        self.branches_table = QTableWidget()
        self.branches_table.setColumnCount(7)
        self.branches_table.setHorizontalHeaderLabels([
            "رمز الفرع", 
            "اسم الفرع", 
            "الموقع", 
            "المحافظة", 
            "عدد الموظفين",
            "الرصيد (ل.س)",  # Syrian Pounds balance
            "الرصيد ($)"      # US Dollars balance
        ])
        self.branches_table.horizontalHeader().setStretchLastSection(True)
        self.branches_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.branches_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.branches_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        
        layout.addWidget(self.branches_table)
        
        # Buttons
        buttons_layout = QHBoxLayout()
        
        add_branch_button = ModernButton("إضافة فرع", color="#2ecc71")
        add_branch_button.clicked.connect(self.add_branch)
        buttons_layout.addWidget(add_branch_button)
        
        edit_branch_button = ModernButton("تعديل الفرع", color="#3498db")
        edit_branch_button.clicked.connect(self.edit_branch)
        buttons_layout.addWidget(edit_branch_button)
        
        # Add new allocation button
        allocate_button = ModernButton("تعيين رصيد", color="#9b59b6")
        allocate_button.clicked.connect(self.allocate_funds)
        buttons_layout.addWidget(allocate_button)
        
        view_fund_history = ModernButton("سجل التمويل", color="#2ecc71")
        view_fund_history.clicked.connect(self.view_fund_history)
        buttons_layout.addWidget(view_fund_history)
        
        delete_branch_button = ModernButton("حذف الفرع", color="#e74c3c")
        delete_branch_button.clicked.connect(self.delete_branch)
        buttons_layout.addWidget(delete_branch_button)
        
        view_branch_button = ModernButton("عرض تفاصيل الفرع", color="#f39c12")
        view_branch_button.clicked.connect(self.view_branch)
        buttons_layout.addWidget(view_branch_button)
        
        refresh_button = ModernButton("تحديث", color="#9b59b6")
        refresh_button.clicked.connect(self.load_branches)
        buttons_layout.addWidget(refresh_button)
        
        layout.addLayout(buttons_layout)
        
        self.branches_tab.setLayout(layout)
        
    def allocate_funds(self):
        """Open dialog to allocate/deduct funds from branch."""
        selected_rows = self.branches_table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "تنبيه", "الرجاء تحديد فرع لتعيين الرصيد")
            return
        
        row = selected_rows[0].row()
        branch_data = self.branches_table.item(row, 0).data(Qt.ItemDataRole.UserRole)
        
        dialog = QDialog(self)
        dialog.setWindowTitle(f"إدارة رصيد الفرع - {branch_data.get('name', '')}")
        dialog.setGeometry(200, 200, 500, 400)
        dialog.setStyleSheet("""
            QDialog {
                background-color: #f5f5f5;
                font-family: Arial;
            }
            QLabel {
                color: #333;
            }
            QRadioButton {
                padding: 8px;
            }
        """)
        
        layout = QVBoxLayout()
        
        # Operation type selection
        type_group = QGroupBox("نوع العملية")
        type_layout = QHBoxLayout()
        
        self.operation_type = QButtonGroup()
        self.add_radio = QRadioButton("إضافة رصيد")
        self.deduct_radio = QRadioButton("خصم رصيد")
        self.add_radio.setChecked(True)
        
        self.operation_type.addButton(self.add_radio)
        self.operation_type.addButton(self.deduct_radio)
        
        type_layout.addWidget(self.add_radio)
        type_layout.addWidget(self.deduct_radio)
        type_group.setLayout(type_layout)
        layout.addWidget(type_group)
        
        # Currency selection
        currency_group = QGroupBox("العملة")
        currency_layout = QHBoxLayout()
        
        self.currency_type = QButtonGroup()
        self.syp_radio = QRadioButton("ليرة سورية (ل.س)")
        self.usd_radio = QRadioButton("دولار أمريكي ($)")
        self.syp_radio.setChecked(True)
        
        self.currency_type.addButton(self.syp_radio)
        self.currency_type.addButton(self.usd_radio)
        
        currency_layout.addWidget(self.syp_radio)
        currency_layout.addWidget(self.usd_radio)
        currency_group.setLayout(currency_layout)
        layout.addWidget(currency_group)
        
        # Amount input
        amount_group = QGroupBox("المبلغ")
        amount_layout = QVBoxLayout()
        
        self.amount_input = QDoubleSpinBox()
        self.amount_input.setRange(0, 100000000)
        self.amount_input.setValue(0)
        self.amount_input.setSingleStep(1000)
        self.amount_input.setButtonSymbols(QAbstractSpinBox.ButtonSymbols.NoButtons)
        
        # Update prefix based on currency selection
        self.syp_radio.toggled.connect(lambda checked: self.amount_input.setPrefix("ل.س " if checked else "$ "))
        self.usd_radio.toggled.connect(lambda checked: self.amount_input.setPrefix("$ " if checked else "ل.س "))
        self.amount_input.setPrefix("ل.س ")  # Initial prefix
        
        amount_layout.addWidget(self.amount_input)
        amount_group.setLayout(amount_layout)
        layout.addWidget(amount_group)
        
        # Description
        desc_group = QGroupBox("الوصف (اختياري)")
        desc_layout = QVBoxLayout()
        self.desc_input = QTextEdit()
        self.desc_input.setMaximumHeight(80)
        desc_layout.addWidget(self.desc_input)
        desc_group.setLayout(desc_layout)
        layout.addWidget(desc_group)
        
        # Buttons
        buttons_layout = QHBoxLayout()
        
        delete_button = ModernButton("حذف الرصيد بالكامل", color="#e67e22")
        delete_button.clicked.connect(lambda: self.delete_allocation(branch_data['id'], dialog))
        
        cancel_button = ModernButton("إلغاء", color="#e74c3c")
        cancel_button.clicked.connect(dialog.reject)
        
        save_button = ModernButton("حفظ التغييرات", color="#2ecc71")
        save_button.clicked.connect(lambda: self.save_allocation(branch_data['id'], dialog))
        
        buttons_layout.addWidget(delete_button)
        buttons_layout.addWidget(cancel_button)
        buttons_layout.addWidget(save_button)
        
        layout.addLayout(buttons_layout)
        
        dialog.setLayout(layout)
        dialog.exec()   
        
    def save_allocation(self, branch_id, dialog):
        """Save allocation/deduction to the branch"""
        try:
            amount = self.amount_input.value()
            operation_type = "allocation" if self.add_radio.isChecked() else "deduction"
            currency = "SYP" if self.syp_radio.isChecked() else "USD"
            description = self.desc_input.toPlainText().strip()
            
            if amount <= 0:
                QMessageBox.warning(self, "خطأ في الإدخال", "يرجى إدخال مبلغ أكبر من الصفر")
                return
                
            headers = {
                "Authorization": f"Bearer {self.token}",
                "Content-Type": "application/json"
            }
            
            data = {
                "amount": amount,
                "type": operation_type,
                "currency": currency,
                "description": description or f"{'إضافة' if operation_type == 'allocation' else 'خصم'} بواسطة المدير"
            }
            
            response = requests.post(
                f"{self.api_url}/branches/{branch_id}/allocate-funds/",
                headers=headers,
                json=data
            )
            
            if response.status_code == 200:
                QMessageBox.information(self, "نجاح", f"تم تحديث الرصيد بنجاح ({currency})")
                self.load_branches()
                dialog.accept()
            else:
                error = response.json().get("detail", "حدث خطأ غير معروف")
                QMessageBox.warning(self, "خطأ", f"فشل في العملية: {error}")
                
        except Exception as e:
            QMessageBox.critical(self, "خطأ فادح", f"حدث خطأ في النظام: {str(e)}")
            
    def delete_allocation(self, branch_id, dialog):
        """Delete all allocated funds for the branch"""
        try:
            confirm = QMessageBox.question(
                self,
                "تأكيد الحذف",
                "هل أنت متأكد من حذف الرصيد بالكامل؟ سيتم ضبط الرصيد على الصفر لكلا العملتين.",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            
            if confirm == QMessageBox.StandardButton.Yes:
                headers = {"Authorization": f"Bearer {self.token}"}
                
                # Delete SYP allocations
                response_syp = requests.delete(
                    f"{self.api_url}/branches/{branch_id}/allocations/?currency=SYP",
                    headers=headers
                )
                
                # Delete USD allocations
                response_usd = requests.delete(
                    f"{self.api_url}/branches/{branch_id}/allocations/?currency=USD",
                    headers=headers
                )
                
                if response_syp.status_code == 200 and response_usd.status_code == 200:
                    QMessageBox.information(self, "نجاح", "تم حذف الرصيد بالكامل لكلا العملتين")
                    self.load_branches()
                    dialog.accept()
                else:
                    error_syp = response_syp.json().get("detail", "") if response_syp.status_code != 200 else ""
                    error_usd = response_usd.json().get("detail", "") if response_usd.status_code != 200 else ""
                    error = error_syp + " " + error_usd
                    QMessageBox.warning(self, "خطأ", f"فشل في الحذف: {error}")
                    
        except Exception as e:
            QMessageBox.critical(self, "خطأ فادح", f"حدث خطأ في النظام: {str(e)}")
    
    def setup_employees_tab(self):
        """Set up the employees tab with proper filtering controls."""
        layout = QVBoxLayout()
        
        # Title
        title = QLabel("إدارة الموظفين")
        title.setFont(QFont("Arial", 18, QFont.Weight.Bold))
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setStyleSheet("color: #2c3e50; margin-bottom: 20px;")
        layout.addWidget(title)
        
        # Filter controls
        filter_group = ModernGroupBox("تصفية الموظفين", "#3498db")
        filter_layout = QGridLayout()
        
        # Branch filter
        branch_label = QLabel("الفرع:")
        self.branch_filter = QComboBox()
        self.branch_filter.setMinimumWidth(250)
        self.load_branches_for_filter()
        self.branch_filter.currentIndexChanged.connect(self.filter_employees)
        
        # Search field
        search_label = QLabel("بحث:")
        self.employee_search = QLineEdit()
        self.employee_search.setPlaceholderText("ابحث باسم الموظف أو المعرف")
        self.employee_search.textChanged.connect(self.filter_employees)
        
        # Add widgets to filter layout
        filter_layout.addWidget(branch_label, 0, 0)
        filter_layout.addWidget(self.branch_filter, 0, 1)
        filter_layout.addWidget(search_label, 1, 0)
        filter_layout.addWidget(self.employee_search, 1, 1)
        
        filter_group.setLayout(filter_layout)
        layout.addWidget(filter_group)
        
        # Employees table
        self.employees_table = QTableWidget()
        self.employees_table.setColumnCount(5)
        self.employees_table.setHorizontalHeaderLabels([
            "اسم المستخدم", "الدور", "الفرع", "تاريخ الإنشاء", "الحالة"
        ])
        self.employees_table.horizontalHeader().setStretchLastSection(True)
        self.employees_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.employees_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.employees_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        
        layout.addWidget(self.employees_table)
        
        # Buttons
        buttons_layout = QHBoxLayout()
        
        add_employee_button = ModernButton("إضافة موظف", color="#2ecc71")
        add_employee_button.clicked.connect(self.add_employee)
        buttons_layout.addWidget(add_employee_button)
        
        edit_employee_button = ModernButton("تعديل الموظف", color="#3498db")
        edit_employee_button.clicked.connect(self.edit_employee)
        buttons_layout.addWidget(edit_employee_button)
        
        delete_employee_button = ModernButton("حذف الموظف", color="#e74c3c")
        delete_employee_button.clicked.connect(self.delete_employee)
        buttons_layout.addWidget(delete_employee_button)
        
        reset_password_button = ModernButton("إعادة تعيين كلمة المرور", color="#f39c12")
        reset_password_button.clicked.connect(self.reset_password)
        buttons_layout.addWidget(reset_password_button)
        
        refresh_button = ModernButton("تحديث", color="#9b59b6")
        refresh_button.clicked.connect(self.load_employees)
        buttons_layout.addWidget(refresh_button)
        
        layout.addLayout(buttons_layout)
        
        self.employee_search.textChanged.connect(self.filter_employees)
        
        self.employees_tab.setLayout(layout)
        
        self.load_employees()  # Initial load
    
    def setup_transactions_tab(self):
        """Set up the transactions tab."""
        layout = QVBoxLayout()
        
        # Title
        title = QLabel("إدارة التحويلات")
        title.setFont(QFont("Arial", 18, QFont.Weight.Bold))
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setStyleSheet("color: #2c3e50; margin-bottom: 20px;")
        layout.addWidget(title)
        
        # Filter
        filter_layout = QHBoxLayout()
        
        filter_branch_label = QLabel("تصفية حسب الفرع:")
        filter_layout.addWidget(filter_branch_label)
        
        self.transaction_branch_filter = QComboBox()
        self.transaction_branch_filter.setMinimumWidth(150)
        self.transaction_branch_filter.currentIndexChanged.connect(self.filter_transactions)
        filter_layout.addWidget(self.transaction_branch_filter)
        
        filter_type_label = QLabel("نوع التصفية:")
        filter_layout.addWidget(filter_type_label)
        
        self.transaction_type_filter = QComboBox()
        self.transaction_type_filter.addItems(["الكل", "الواردة", "الصادرة", "متعلقة بالفرع"])
        self.transaction_type_filter.setMinimumWidth(150)
        self.transaction_type_filter.currentIndexChanged.connect(self.filter_transactions)
        filter_layout.addWidget(self.transaction_type_filter)
        
        # Add status filter using the provided statuses list
        status_filter_label = QLabel("تصفية حسب الحالة:")
        filter_layout.addWidget(status_filter_label)
        
        self.status_filter = QComboBox()
        # Add "All" option first
        self.status_filter.addItem("الكل", "all")
        
        # Add status options from the provided list
        statuses = [
            ("قيد الانتظار", "pending"),
            ("قيد المعالجة", "processing"),
            ("مكتمل", "completed"),
            ("ملغي", "cancelled"),
            ("مرفوض", "rejected"),
            ("معلق", "on_hold")
        ]
        
        for status_arabic, status_code in statuses:
            self.status_filter.addItem(status_arabic, status_code)
            
        self.status_filter.setMinimumWidth(150)
        self.status_filter.currentIndexChanged.connect(self.filter_transactions)
        filter_layout.addWidget(self.status_filter)
        
        filter_layout.addStretch()
        
        search_button = ModernButton("بحث", color="#3498db")
        search_button.clicked.connect(self.search_transaction)
        filter_layout.addWidget(search_button)
        
        layout.addLayout(filter_layout)
        
        # Transactions table
        self.transactions_table = QTableWidget()
        self.transactions_table.setColumnCount(11)
        self.transactions_table.setHorizontalHeaderLabels([
            "النوع", "رقم التحويل", "المرسل", "المستلم", "المبلغ", "التاريخ", "الحالة", 
            "الفرع المرسل", "اتجاه الصادر", "الفرع المستلم", "اتجاه الوارد", "اسم الموظف"
        ])
        self.transactions_table.horizontalHeader().setStretchLastSection(True)
        self.transactions_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.transactions_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.transactions_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.transactions_table.setStyleSheet("""
            QTableWidget {
                border: none;
                background-color: white;
                gridline-color: #ddd;
            }
            QHeaderView::section {
                background-color: #2c3e50;
                color: white;
                padding: 8px;
                border: 1px solid #1a2530;
                font-weight: bold;
            }
            QTableWidget::item {
                padding: 5px;
            }
            QTableWidget::item:selected {
                background-color: #3498db;
                color: white;
            }
        """)
        
        layout.addWidget(self.transactions_table)
        
        # Add pagination controls
        pagination_layout = QHBoxLayout()
        
        self.prev_trans_button = ModernButton("السابق", color="#3498db")
        self.prev_trans_button.clicked.connect(self.prev_trans_page)
        pagination_layout.addWidget(self.prev_trans_button)
        
        self.trans_page_label = QLabel("الصفحة: 1")
        pagination_layout.addWidget(self.trans_page_label)
        
        self.next_trans_button = ModernButton("التالي", color="#3498db")
        self.next_trans_button.clicked.connect(self.next_trans_page)
        pagination_layout.addWidget(self.next_trans_button)
        
        layout.addLayout(pagination_layout)
        
        # Buttons
        buttons_layout = QHBoxLayout()
        
        view_transaction_button = ModernButton("عرض التفاصيل", color="#3498db")
        view_transaction_button.clicked.connect(self.view_transaction)
        buttons_layout.addWidget(view_transaction_button)
        
        update_status_button = ModernButton("تحديث الحالة", color="#f39c12")
        update_status_button.clicked.connect(self.update_transaction_status)
        buttons_layout.addWidget(update_status_button)
        
        print_receipt_button = ModernButton("طباعة الإيصال", color="#2ecc71")
        print_receipt_button.clicked.connect(self.print_receipt)
        buttons_layout.addWidget(print_receipt_button)
        
        refresh_button = ModernButton("تحديث", color="#9b59b6")
        refresh_button.clicked.connect(self.load_transactions)
        buttons_layout.addWidget(refresh_button)
        
        layout.addLayout(buttons_layout)
        
        self.transactions_tab.setLayout(layout)
    
    def setup_reports_tab(self):
        """Set up the reports tab."""
        layout = QVBoxLayout()
        
        # Title
        title = QLabel("التقارير")
        title.setFont(QFont("Arial", 18, QFont.Weight.Bold))
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setStyleSheet("color: #2c3e50; margin-bottom: 20px;")
        layout.addWidget(title)
        
        # Report options
        options_group = ModernGroupBox("خيارات التقرير", "#3498db")
        options_layout = QGridLayout()
        
        # Report type
        report_type_label = QLabel("نوع التقرير:")
        options_layout.addWidget(report_type_label, 0, 0)
        
        self.report_type = QComboBox()
        self.report_type.addItems(["تقرير التحويلات", "تقرير الفروع", "تقرير الموظفين"])
        options_layout.addWidget(self.report_type, 0, 1)
        
        # Date range
        date_range_label = QLabel("نطاق التاريخ:")
        options_layout.addWidget(date_range_label, 1, 0)
        
        date_range_layout = QHBoxLayout()
        
        from_date_label = QLabel("من:")
        date_range_layout.addWidget(from_date_label)
        
        self.from_date = QDateEdit()
        self.from_date.setCalendarPopup(True)
        self.from_date.setDate(QDate.currentDate().addDays(-30))  # Last 30 days
        date_range_layout.addWidget(self.from_date)
        
        to_date_label = QLabel("إلى:")
        date_range_layout.addWidget(to_date_label)
        
        self.to_date = QDateEdit()
        self.to_date.setCalendarPopup(True)
        self.to_date.setDate(QDate.currentDate())
        date_range_layout.addWidget(self.to_date)
        
        options_layout.addLayout(date_range_layout, 1, 1)
        
        # Branch filter
        branch_filter_label = QLabel("الفرع:")
        options_layout.addWidget(branch_filter_label, 2, 0)
        
        self.report_branch_filter = QComboBox()
        options_layout.addWidget(self.report_branch_filter, 2, 1)
        
        # Generate button
        generate_button = ModernButton("إنشاء التقرير", color="#2ecc71")
        generate_button.clicked.connect(self.generate_report)
        options_layout.addWidget(generate_button, 3, 0, 1, 2)
        
        options_group.setLayout(options_layout)
        layout.addWidget(options_group)
        
        # Report preview
        preview_group = ModernGroupBox("معاينة التقرير", "#e74c3c")
        preview_layout = QVBoxLayout()
        
        self.report_table = QTableWidget()
        preview_layout.addWidget(self.report_table)
        
        # Export buttons
        export_layout = QHBoxLayout()
        
        export_pdf_button = ModernButton("تصدير PDF", color="#3498db")
        export_pdf_button.clicked.connect(self.export_pdf)
        export_layout.addWidget(export_pdf_button)
        
        export_excel_button = ModernButton("تصدير Excel", color="#f39c12")
        export_excel_button.clicked.connect(self.export_excel)
        export_layout.addWidget(export_excel_button)
        
        export_print_button = ModernButton("طباعة", color="#9b59b6")
        export_print_button.clicked.connect(self.print_report)
        export_layout.addWidget(export_print_button)
        
        preview_layout.addLayout(export_layout)
        preview_group.setLayout(preview_layout)
        layout.addWidget(preview_group)
        
        self.reports_tab.setLayout(layout)
        # We'll load branches in the init method to ensure all tabs are properly populated
    
    def setup_settings_tab(self):
        """Set up the settings tab."""
        layout = QVBoxLayout()
        
        # System settings
        settings_group = ModernGroupBox("إعدادات النظام", "#3498db")
        settings_layout = QFormLayout()
        
        self.system_name_input = QLineEdit("نظام التحويلات المالية الداخلي")
        settings_layout.addRow("اسم النظام:", self.system_name_input)
        
        self.company_name_input = QLineEdit("شركة التحويلات المالية")
        settings_layout.addRow("اسم الشركة:", self.company_name_input)
        
        self.admin_email_input = QLineEdit("admin@example.com")
        settings_layout.addRow("البريد الإلكتروني للمسؤول:", self.admin_email_input)
        
        self.currency_input = QComboBox()
        self.currency_input.addItems(["ليرة سورية", "دولار أمريكي", "يورو"])
        settings_layout.addRow("العملة الافتراضية:", self.currency_input)
        
        self.language_input = QComboBox()
        self.language_input.addItems(["العربية", "English"])
        settings_layout.addRow("اللغة:", self.language_input)
        
        self.theme_input = QComboBox()
        self.theme_input.addItems(["فاتح", "داكن", "أزرق"])
        settings_layout.addRow("السمة:", self.theme_input)
        
        save_settings_button = ModernButton("حفظ الإعدادات", color="#2ecc71")
        save_settings_button.clicked.connect(self.save_settings)
        settings_layout.addRow("", save_settings_button)
        
        settings_group.setLayout(settings_layout)
        layout.addWidget(settings_group)
        
        # User settings
        user_settings_group = ModernGroupBox("إعدادات المستخدم", "#e74c3c")
        user_settings_layout = QFormLayout()
        
        self.username_input = QLineEdit("admin")
        self.username_input.setReadOnly(True)
        self.username_input.setStyleSheet("background-color: #f0f0f0;")
        user_settings_layout.addRow("اسم المستخدم:", self.username_input)
        
        self.old_password_input = QLineEdit()
        self.old_password_input.setEchoMode(QLineEdit.EchoMode.Password)
        user_settings_layout.addRow("كلمة المرور الحالية:", self.old_password_input)
        
        self.new_password_input = QLineEdit()
        self.new_password_input.setEchoMode(QLineEdit.EchoMode.Password)
        user_settings_layout.addRow("كلمة المرور الجديدة:", self.new_password_input)
        
        self.confirm_password_input = QLineEdit()
        self.confirm_password_input.setEchoMode(QLineEdit.EchoMode.Password)
        user_settings_layout.addRow("تأكيد كلمة المرور:", self.confirm_password_input)
        
        change_password_button = ModernButton("تغيير كلمة المرور", color="#f39c12")
        change_password_button.clicked.connect(self.change_password)
        user_settings_layout.addRow("", change_password_button)
        
        user_settings_group.setLayout(user_settings_layout)
        layout.addWidget(user_settings_group)
        
        # Backup and restore
        backup_group = ModernGroupBox("النسخ الاحتياطي واستعادة البيانات", "#9b59b6")
        backup_layout = QVBoxLayout()
        
        backup_button = ModernButton("إنشاء نسخة احتياطية", color="#3498db")
        backup_button.clicked.connect(self.create_backup)
        backup_layout.addWidget(backup_button)
        
        restore_button = ModernButton("استعادة من نسخة احتياطية", color="#e74c3c")
        restore_button.clicked.connect(self.restore_backup)
        backup_layout.addWidget(restore_button)
        
        backup_group.setLayout(backup_layout)
        layout.addWidget(backup_group)
        
        self.settings_tab.setLayout(layout)
    
    def load_dashboard_data(self):
        """Load data for the dashboard."""
        try:
            headers = {"Authorization": f"Bearer {self.token}"} if self.token else {}
            
            # Load all dashboard components
            self.load_branch_stats()  # This now updates branches count
            
            # Get employee stats
            response = requests.get(f"{self.api_url}/users/stats/", headers=headers)
            if response.status_code == 200:
                data = response.json()
                self.employees_count.setText(str(data.get("total", 0)))
            
            # Get transaction stats
            response = requests.get(f"{self.api_url}/transactions/stats/", headers=headers)
            if response.status_code == 200:
                data = response.json()
                self.transactions_count.setText(str(data.get("total", 0)))
                self.amount_total.setText(f"{data.get('total_amount', 0):,.2f}")
            
            # Load recent transactions
            self.load_recent_transactions()
            
        except Exception as e:
            print(f"Error loading dashboard data: {e}")
            QMessageBox.warning(self, "خطأ", f"تعذر تحميل بيانات لوحة المعلومات: {str(e)}")
    
    def load_recent_transactions(self):
        """Load recent transactions with directional branch information"""
        try:
            headers = {"Authorization": f"Bearer {self.token}"}
            response = requests.get(f"{self.api_url}/transactions/", headers=headers)
            
            if response.status_code == 200:
                data = response.json()
                all_transactions = data.get("transactions", [])
                
                # Client-side pagination
                self.total_pages = (len(all_transactions) + self.per_page - 1) // self.per_page
                start_index = (self.current_page - 1) * self.per_page
                end_index = start_index + self.per_page
                transactions = all_transactions[start_index:end_index]
                
                # Create branch name mapping
                if not hasattr(self, 'branch_id_to_name'):
                    self.branch_id_to_name = {}
                    branches_response = requests.get(f"{self.api_url}/branches/", headers=headers)
                    if branches_response.status_code == 200:
                        branches = branches_response.json().get("branches", [])
                        self.branch_id_to_name = {b["id"]: b["name"] for b in branches}
                
                # Set column count and headers (if not already done)
                self.recent_transactions_table.setColumnCount(12)
                self.recent_transactions_table.setHorizontalHeaderLabels([
                    "النوع", "رقم التحويل", "المرسل", "المستلم", "المبلغ", 
                    "التاريخ", "الحالة", "الفرع المرسل", "اتجاه الصادر",
                    "الفرع المستلم", "اتجاه الوارد", "اسم الموظف"
                ])
                
                # Clear the table first
                self.recent_transactions_table.setRowCount(0)
                
                self.recent_transactions_table.setRowCount(len(transactions))
                
                for row, transaction in enumerate(transactions):
                    # Transaction Type
                    type_item = self.create_transaction_type_item(transaction)
                    self.recent_transactions_table.setItem(row, 0, type_item)
                    
                    # Transaction ID
                    trans_id = str(transaction.get("id", ""))
                    id_item = QTableWidgetItem(trans_id[:8] + "..." if len(trans_id) > 8 else trans_id)
                    id_item.setToolTip(trans_id)
                    self.recent_transactions_table.setItem(row, 1, id_item)
                    
                    # Sender/Receiver
                    self.recent_transactions_table.setItem(row, 2, QTableWidgetItem(transaction.get("sender", "")))
                    self.recent_transactions_table.setItem(row, 3, QTableWidgetItem(transaction.get("receiver", "")))
                    
                    # Amount
                    amount = transaction.get("amount", 0)
                    currency = transaction.get("currency", "ليرة سورية")
                    amount_item = QTableWidgetItem(f"{amount:,.2f} {currency}")
                    amount_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                    self.recent_transactions_table.setItem(row, 4, amount_item)
                    
                    # Date
                    date_str = transaction.get("date", "")
                    self.recent_transactions_table.setItem(row, 5, QTableWidgetItem(date_str))
                    
                    # Status
                    status = transaction.get("status", "").lower()
                    status_ar = self.get_status_arabic(status)
                    status_item = QTableWidgetItem(status_ar)
                    status_item.setBackground(self.get_status_color(status))
                    status_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.recent_transactions_table.setItem(row, 6, status_item)
                    
                    # Sending Branch and Direction
                    branch_id = transaction.get("branch_id")
                    sending_branch = self.branch_id_to_name.get(branch_id, f"الفرع {branch_id}" if branch_id else "غير معروف")
                    self.recent_transactions_table.setItem(row, 7, QTableWidgetItem(sending_branch))
                    
                    # Outgoing Direction Indicator
                    outgoing_direction = QTableWidgetItem("↑" if branch_id else "")
                    outgoing_direction.setForeground(QColor(0, 150, 0))  # Green color
                    outgoing_direction.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.recent_transactions_table.setItem(row, 8, outgoing_direction)
                    
                    # Receiving Branch and Direction
                    dest_branch_id = transaction.get("destination_branch_id")
                    receiving_branch = self.branch_id_to_name.get(dest_branch_id, f"الفرع {dest_branch_id}" if dest_branch_id else "غير معروف")
                    self.recent_transactions_table.setItem(row, 9, QTableWidgetItem(receiving_branch))
                    
                    # Incoming Direction Indicator
                    incoming_direction = QTableWidgetItem("↓" if dest_branch_id else "")
                    incoming_direction.setForeground(QColor(150, 0, 0))  # Red color
                    incoming_direction.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.recent_transactions_table.setItem(row, 10, incoming_direction)
                    
                    # Employee
                    self.recent_transactions_table.setItem(row, 11, QTableWidgetItem(transaction.get("employee_name", "")))

                self.update_pagination_controls()
                
        except Exception as e:
            print(f"Error loading recent transactions: {e}")
            self.statusBar().showMessage(f"خطأ في تحميل التحويلات: {str(e)}", 5000)

            
    def get_status_arabic(self, status):
        """Convert status to Arabic."""
        status_map = {
            "processing": "قيد المعالجة",
            "completed": "مكتمل",
            "cancelled": "ملغي",
            "rejected": "مرفوض",
            "on_hold": "معلق"
        }
        return status_map.get(status, status)

    def get_status_color(self, status):
        """Get color for status."""
        status_colors = {
            "processing": QColor(200, 200, 255),  # Light blue
            "completed": QColor(200, 255, 200),   # Light green
            "cancelled": QColor(255, 200, 200),   # Light red
            "rejected": QColor(255, 150, 150),    # Darker red
            "on_hold": QColor(255, 200, 150)      # Light orange
        }
        return status_colors.get(status, QColor(255, 255, 255))

    def update_pagination_controls(self):
        """Update pagination controls."""
        self.page_label.setText(f"الصفحة: {self.current_page}/{self.total_pages}")
        self.prev_button.setEnabled(self.current_page > 1)
        self.next_button.setEnabled(self.current_page < self.total_pages)

    def prev_page(self):
        """Navigate to previous page."""
        if self.current_page > 1:
            self.current_page -= 1
            self.load_recent_transactions()

    def next_page(self):
        """Navigate to next page."""
        if self.current_page < self.total_pages:
            self.current_page += 1
            self.load_recent_transactions()     
    
    def load_branches(self):
        """Load branches data."""
        try:
            headers = {"Authorization": f"Bearer {self.token}"} if self.token else {}
            
            response = requests.get(f"{self.api_url}/branches/", headers=headers)
            
            if response.status_code == 200:
                try:
                    response_data = response.json()
                    branches = response_data.get("branches", [])
                    
                    if not isinstance(branches, list):
                        QMessageBox.warning(self, "خطأ", "تنسيق بيانات الفروع غير صحيح")
                        return
                    
                    # Debug print to verify branches data
                    print(f"Received {len(branches)} branches")
                    for branch in branches:
                        print(f"Branch data: {branch}")
                    
                    self.branches_table.setRowCount(len(branches))
                    
                    for i, branch in enumerate(branches):
                        if not isinstance(branch, dict):
                            print(f"Debug: Invalid branch data at index {i}: {branch}")
                            continue
                            
                        # Column 0: Branch ID
                        self.branches_table.setItem(i, 0, QTableWidgetItem(branch.get("branch_id", "")))
                        
                        # Column 1: Branch Name
                        self.branches_table.setItem(i, 1, QTableWidgetItem(branch.get("name", "")))
                        
                        # Column 2: Location
                        self.branches_table.setItem(i, 2, QTableWidgetItem(branch.get("location", "")))
                        
                        # Column 3: Governorate
                        self.branches_table.setItem(i, 3, QTableWidgetItem(branch.get("governorate", "")))
                        
                        # Column 4: Number of Employees (Corrected)
                        branch_id = branch.get("id")
                        emp_count = 0
                        if branch_id:
                            try:
                                emp_response = requests.get(
                                    f"{self.api_url}/branches/{branch_id}/employees/stats/", 
                                    headers=headers
                                )
                                if emp_response.status_code == 200:
                                    emp_data = emp_response.json()
                                    emp_count = emp_data.get("total", 0)
                            except Exception as emp_e:
                                print(f"Debug: Error getting employee count: {emp_e}")
                        self.branches_table.setItem(i, 4, QTableWidgetItem(str(emp_count)))
                        
                        # Column 5: Syrian Pounds Balance
                        syp_allocated = branch.get("allocated_amount_syp", branch.get("allocated_amount", 0))
                        syp_amount_item = QTableWidgetItem(f"{syp_allocated:,.2f} ل.س")
                        syp_amount_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                        self.branches_table.setItem(i, 5, syp_amount_item)
                        
                        # Column 6: US Dollars Balance
                        usd_allocated = branch.get("allocated_amount_usd", 0)
                        usd_amount_item = QTableWidgetItem(f"{usd_allocated:,.2f} $")
                        usd_amount_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                        self.branches_table.setItem(i, 6, usd_amount_item)
                        
                        # Debug print to verify data
                        print(f"Branch {branch.get('name')}: SYP={syp_allocated}, USD={usd_allocated}")
                        
                        # Store branch data
                        self.branches_table.item(i, 0).setData(Qt.ItemDataRole.UserRole, branch)
                        
                except ValueError as json_e:
                    print(f"Debug: JSON parse error: {json_e}")
                    QMessageBox.warning(self, "خطأ", "تعذر تحليل استجابة الخادم")
                    
            else:
                error_msg = f"فشل تحميل الفروع: رمز الحالة {response.status_code}"
                try:
                    error_data = response.json()
                    if "detail" in error_data:
                        error_msg = error_data["detail"]
                except:
                    pass
                QMessageBox.warning(self, "خطأ", error_msg)
                
        except Exception as e:
            print(f"Debug: Error in load_branches: {e}")
            QMessageBox.warning(self, "خطأ", f"تعذر تحميل الفروع: {str(e)}")
            
    def view_fund_history(self):
        selected_rows = self.branches_table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "تنبيه", "الرجاء تحديد فرع لعرض السجل")
            return
        
        row = selected_rows[0].row()
        branch_data = self.branches_table.item(row, 0).data(Qt.ItemDataRole.UserRole)
        
        dialog = QDialog(self)
        dialog.setWindowTitle(f"سجل التمويل - {branch_data.get('name', '')}")
        dialog.setGeometry(100, 100, 600, 400)
        
        layout = QVBoxLayout()
        
        table = QTableWidget()
        table.setColumnCount(4)
        table.setHorizontalHeaderLabels(["التاريخ", "النوع", "المبلغ", "الوصف"])
        
        try:
            headers = {"Authorization": f"Bearer {self.token}"}
            response = requests.get(
                f"{self.api_url}/branches/{branch_data['id']}/funds-history",
                headers=headers
            )
            
            if response.status_code == 200:
                history = response.json()
                table.setRowCount(len(history))
                
                for i, record in enumerate(history):
                    # تعريب نوع الحركة
                    type_ar = {
                        "allocation": "إيداع",
                        "deduction": "خصم",
                        "refund": "استرداد"
                    }.get(record.get('type', ''), record.get('type', ''))
                    
                    # التاريخ
                    date_item = QTableWidgetItem(record.get('date', 'غير معروف'))
                    
                    # المبلغ
                    amount = record.get('amount', 0)
                    amount_item = QTableWidgetItem(f"{amount:,.2f} ل.س")
                    amount_item.setTextAlignment(Qt.AlignmentFlag.AlignRight)
                    
                    table.setItem(i, 0, date_item)
                    table.setItem(i, 1, QTableWidgetItem(type_ar))
                    table.setItem(i, 2, amount_item)
                    table.setItem(i, 3, QTableWidgetItem(record.get('description', '')))
        
        except Exception as e:
            QMessageBox.warning(dialog, "خطأ", f"خطأ في جلب البيانات: {str(e)}")
        
        layout.addWidget(table)
        dialog.setLayout(layout)
        dialog.exec()     
    
    def load_employees(self, branch_id=None):
        """Load employees data."""
        try:
            headers = {"Authorization": f"Bearer {self.token}"} if self.token else {}
            
            # Use /users/ endpoint instead of /employees/ to include branch managers
            url = f"{self.api_url}/users/"
            if branch_id:
                url += f"?branch_id={branch_id}"
            
            response = requests.get(url, headers=headers)
            
            if response.status_code == 200:
                # For /users/ endpoint, the response is wrapped in a "users" key
                employees = response.json().get("users", [])
                self.employees_table.setRowCount(len(employees))
                
                for i, employee in enumerate(employees):
                    self.employees_table.setItem(i, 0, QTableWidgetItem(employee.get("username", "")))
                    
                    # Map role to Arabic
                    role = employee.get("role", "")
                    role_arabic = "موظف"
                    if role == "director":
                        role_arabic = "مدير"
                    elif role == "branch_manager":
                        role_arabic = "مدير فرع"
                    
                    self.employees_table.setItem(i, 1, QTableWidgetItem(role_arabic))
                    
                    # Get branch name
                    branch_id = employee.get("branch_id")
                    branch_name = "غير محدد"
                    if branch_id:
                        try:
                            branch_response = requests.get(
                                f"{self.api_url}/branches/{branch_id}", 
                                headers=headers
                            )
                            if branch_response.status_code == 200:
                                branch_data = branch_response.json()
                                branch_name = branch_data.get("name", "غير محدد")
                        except:
                            pass
                    
                    self.employees_table.setItem(i, 2, QTableWidgetItem(branch_name))
                    self.employees_table.setItem(i, 3, QTableWidgetItem(employee.get("created_at", "")))
                    
                    # Status (always active for now)
                    status_item = QTableWidgetItem("نشط")
                    status_item.setForeground(QColor("#27ae60"))
                    self.employees_table.setItem(i, 4, status_item)
                    
                    # Store the employee data in the first cell for later use
                    self.employees_table.item(i, 0).setData(Qt.ItemDataRole.UserRole, employee)
            else:
                QMessageBox.warning(self, "خطأ", f"فشل تحميل الموظفين: رمز الحالة {response.status_code}")
        except Exception as e:
            print(f"Error loading employees: {e}")
            QMessageBox.warning(self, "خطأ", f"تعذر تحميل الموظفين: {str(e)}")
            
    def load_branch_stats(self):
        try:
            headers = {"Authorization": f"Bearer {self.token}"}
            response = requests.get(f"{self.api_url}/branches/stats/", headers=headers)
            
            if response.status_code == 200:
                data = response.json()
                branches = data.get("branches", [])
                
                # Update branches count in dashboard
                self.branches_count.setText(str(data.get("total", 0)))
                
                if not branches:
                    self.transfers_bars.setText("لا توجد بيانات متاحة")
                    self.amounts_bars.setText("لا توجد بيانات متاحة")
                    return
                
                # Get top 3 branches for transfers
                top_transfers = sorted(
                    branches, 
                    key=lambda x: x['transactions_count'], 
                    reverse=True
                )[:3]

                # Get top 3 branches for amounts
                top_amounts = sorted(
                    branches, 
                    key=lambda x: x['total_amount'], 
                    reverse=True
                )[:3]

                # Prepare transfers HTML
                transfers_html = ["<table width='100%'>"]
                if top_transfers:
                    max_transfers = max(t['transactions_count'] for t in top_transfers)
                    for branch in top_transfers:
                        width = (branch['transactions_count'] / max_transfers * 80) if max_transfers != 0 else 0
                        transfers_html.append(
                            f"<tr>"
                            f"<td width='30%' align='right'>{branch['name']}</td>"
                            f"<td width='60%'><div style='background: #3498db; height: 20px; width: {width}%; "
                            f"border-radius: 10px; margin: 2px;'></div></td>"
                            f"<td width='10%' align='left'>{branch['transactions_count']}</td>"
                            f"</tr>"
                        )
                transfers_html.append("</table>")
                
                # Prepare amounts HTML
                amounts_html = ["<table width='100%'>"]
                if top_amounts:
                    max_amount = max(a['total_amount'] for a in top_amounts)
                    for branch in top_amounts:
                        width = (branch['total_amount'] / max_amount * 80) if max_amount != 0 else 0
                        amounts_html.append(
                            f"<tr>"
                            f"<td width='30%' align='right'>{branch['name']}</td>"
                            f"<td width='60%'><div style='background: #e74c3c; height: 20px; width: {width}%; "
                            f"border-radius: 10px; margin: 2px;'></div></td>"
                            f"<td width='10%' align='left'>{branch['total_amount']:,.2f}</td>"
                            f"</tr>"
                        )
                amounts_html.append("</table>")

                self.transfers_bars.setText("".join(transfers_html))
                self.amounts_bars.setText("".join(amounts_html))
            else:
                self.transfers_bars.setText("حدث خطأ في جلب البيانات")
                self.amounts_bars.setText("حدث خطأ في جلب البيانات")

        except Exception as e:
            print(f"Error loading branch stats: {e}")
            self.transfers_bars.setText("تعذر تحميل بيانات التحويلات")
            self.amounts_bars.setText("تعذر تحميل بيانات المبالغ")
        
    
    def load_transactions(self, branch_id=None, filter_type="all", status=None):
        """Load transactions data with pagination and matching recent transactions style"""
        try:
            headers = {"Authorization": f"Bearer {self.token}"} if self.token else {}
            
            url = f"{self.api_url}/transactions/"
            params = {}
            
            # Handle branch filtering
            if branch_id and isinstance(branch_id, (int, str)) and branch_id != self.api_url:
                try:
                    if isinstance(branch_id, str) and branch_id.isdigit():
                        branch_id = int(branch_id)
                    params["branch_id"] = branch_id
                except:
                    print(f"Invalid branch_id: {branch_id}")
            
            # Handle filter type
            if filter_type != "all":
                params["filter_type"] = filter_type
                
            # Handle status filtering
            if status and status != "all":
                params["status"] = status
            
            # Fetch all transactions
            response = requests.get(url, headers=headers, params=params)
            
            if response.status_code == 200:
                data = response.json()
                all_transactions = data.get("transactions", [])
                
                # Setup pagination
                self.transactions_per_page = 15
                self.total_pages_transactions = (len(all_transactions) + self.transactions_per_page - 1) // self.transactions_per_page
                start_index = (self.current_page_transactions - 1) * self.transactions_per_page
                end_index = start_index + self.transactions_per_page
                transactions = all_transactions[start_index:end_index]
                
                # Initialize table structure
                self.transactions_table.setColumnCount(12)
                self.transactions_table.setHorizontalHeaderLabels([
                    "النوع", "رقم التحويل", "المرسل", "المستلم", "المبلغ", 
                    "التاريخ", "الحالة", "الفرع المرسل", "اتجاه الصادر",
                    "الفرع المستلم", "اتجاه الوارد", "اسم الموظف"
                ])
                self.transactions_table.setRowCount(len(transactions))
                
                # Load branch names if not cached
                if not hasattr(self, 'branch_id_to_name'):
                    self.branch_id_to_name = {}
                    branches_response = requests.get(f"{self.api_url}/branches/", headers=headers)
                    if branches_response.status_code == 200:
                        branches = branches_response.json().get("branches", [])
                        self.branch_id_to_name = {b["id"]: b["name"] for b in branches}
                
                # Populate table rows
                for row, transaction in enumerate(transactions):
                    # Transaction Type
                    type_item = self.create_transaction_type_item(transaction)
                    self.transactions_table.setItem(row, 0, type_item)
                    
                    # Transaction ID
                    trans_id = str(transaction.get("id", ""))
                    id_item = QTableWidgetItem(trans_id[:8] + "..." if len(trans_id) > 8 else trans_id)
                    id_item.setToolTip(trans_id)
                    self.transactions_table.setItem(row, 1, id_item)
                    
                    # Sender/Receiver
                    self.transactions_table.setItem(row, 2, QTableWidgetItem(transaction.get("sender", "")))
                    self.transactions_table.setItem(row, 3, QTableWidgetItem(transaction.get("receiver", "")))
                    
                    # Amount
                    amount = transaction.get("amount", 0)
                    currency = transaction.get("currency", "ليرة سورية")
                    amount_item = QTableWidgetItem(f"{amount:,.2f} {currency}")
                    amount_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                    self.transactions_table.setItem(row, 4, amount_item)
                    
                    # Date
                    date_str = transaction.get("date", "")
                    self.transactions_table.setItem(row, 5, QTableWidgetItem(date_str))
                    
                    # Status
                    status = transaction.get("status", "").lower()
                    status_ar = self.get_status_arabic(status)
                    status_item = QTableWidgetItem(status_ar)
                    status_item.setBackground(self.get_status_color(status))
                    status_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.transactions_table.setItem(row, 6, status_item)
                    
                    # Branch information and directions
                    branch_id = transaction.get("branch_id")
                    dest_branch_id = transaction.get("destination_branch_id")
                    
                    # Sending Branch
                    sending_branch = self.branch_id_to_name.get(branch_id, f"الفرع {branch_id}" if branch_id else "غير معروف")
                    self.transactions_table.setItem(row, 7, QTableWidgetItem(sending_branch))
                    
                    # Outgoing Direction
                    outgoing_direction = QTableWidgetItem("↑" if branch_id else "")
                    outgoing_direction.setForeground(QColor(0, 150, 0))
                    outgoing_direction.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.transactions_table.setItem(row, 8, outgoing_direction)
                    
                    # Receiving Branch
                    receiving_branch = self.branch_id_to_name.get(dest_branch_id, f"الفرع {dest_branch_id}" if dest_branch_id else "غير معروف")
                    self.transactions_table.setItem(row, 9, QTableWidgetItem(receiving_branch))
                    
                    # Incoming Direction
                    incoming_direction = QTableWidgetItem("↓" if dest_branch_id else "")
                    incoming_direction.setForeground(QColor(150, 0, 0))
                    incoming_direction.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.transactions_table.setItem(row, 10, incoming_direction)
                    
                    # Employee
                    self.transactions_table.setItem(row, 11, QTableWidgetItem(transaction.get("employee_name", "")))
                    
                    # Store transaction data
                    self.transactions_table.item(row, 0).setData(Qt.ItemDataRole.UserRole, transaction)
                
                self.update_trans_pagination_controls()
                
            else:
                QMessageBox.warning(self, "خطأ", f"فشل تحميل التحويلات: رمز الحالة {response.status_code}")
        
        except Exception as e:
            print(f"Error loading transactions: {e}")
            QMessageBox.warning(self, "خطأ", f"تعذر تحميل التحويلات: {str(e)}")
    # Add these helper methods to the class
    def update_trans_pagination_controls(self):
        """Update transactions pagination controls."""
        self.trans_page_label.setText(f"الصفحة: {self.current_page_transactions}/{self.total_pages_transactions}")
        self.prev_trans_button.setEnabled(self.current_page_transactions > 1)
        self.next_trans_button.setEnabled(self.current_page_transactions < self.total_pages_transactions)

    def prev_trans_page(self):
        """Navigate to previous transactions page."""
        if self.current_page_transactions > 1:
            self.current_page_transactions -= 1
            self.load_transactions()

    def next_trans_page(self):
        """Navigate to next transactions page."""
        if self.current_page_transactions < self.total_pages_transactions:
            self.current_page_transactions += 1
            self.load_transactions()
    
    def load_branches_for_filter(self):
        """Load branches for filter dropdowns in all tabs."""
        try:
            headers = {"Authorization": f"Bearer {self.token}"} if self.token else {}
            response = requests.get(f"{self.api_url}/branches/", headers=headers)
            
            if response.status_code == 200:
                response_data = response.json()
                branches = response_data.get("branches", [])
                
                # Clear and populate branch filter in Employees tab
                if hasattr(self, 'branch_filter'):
                    self.branch_filter.clear()
                    self.branch_filter.addItem("جميع الفروع", None)  # إضافة خيار "الكل"
                    
                    for branch in branches:
                        branch_id = branch.get("id")
                        branch_name = branch.get("name", "")
                        if branch_id and branch_name:
                            # تخزين الـ ID كبيانات مرفقة مع الاسم
                            self.branch_filter.addItem(branch_name, branch_id)
                
                # Clear and populate branch filter in Transactions tab
                if hasattr(self, 'transaction_branch_filter'):
                    self.transaction_branch_filter.clear()
                    self.transaction_branch_filter.addItem("جميع الفروع", None)
                    
                    for branch in branches:
                        branch_id = branch.get("id")
                        branch_name = branch.get("name", "")
                        if branch_id and branch_name:
                            self.transaction_branch_filter.addItem(branch_name, branch_id)
                
                # Clear and populate branch filter in Reports tab
                if hasattr(self, 'report_branch_filter'):
                    self.report_branch_filter.clear()
                    self.report_branch_filter.addItem("جميع الفروع", None)
                    
                    for branch in branches:
                        branch_id = branch.get("id")
                        branch_name = branch.get("name", "")
                        if branch_id and branch_name:
                            self.report_branch_filter.addItem(branch_name, branch_id)
                
                # Create branch_id_to_name mapping for future use
                self.branch_id_to_name = {}
                for branch in branches:
                    branch_id = branch.get("id")
                    branch_name = branch.get("name", "")
                    if branch_id and branch_name:
                        self.branch_id_to_name[branch_id] = branch_name
                            
        except Exception as e:
            print(f"Error loading branches: {e}")
            QMessageBox.warning(self, "خطأ", "تعذر تحميل قائمة الفروع")
    
    def filter_employees(self):
        try:
            # مسح الجدول قبل التحميل
            self.employees_table.setRowCount(0)
            
            # استخراج معايير التصفية
            branch_id = self.branch_filter.currentData()
            search_text = self.employee_search.text().strip().lower()

            # إعداد بارامترات الطلب
            params = {}
            if branch_id and branch_id != self.api_url and branch_id != "":  # تجنب إرسال branch_id إذا كان None (جميع الفروع)
                params["branch_id"] = branch_id

            # إرسال الطلب مع البارامترات
            headers = {"Authorization": f"Bearer {self.token}"} if self.token else {}
            response = requests.get(f"{self.api_url}/users/", headers=headers, params=params)
            
            if response.status_code == 200:
                employees = response.json().get("users", [])
                
                # تطبيق البحث المحلي إذا كان هناك نص بحث
                if search_text:
                    filtered_employees = []
                    for emp in employees:
                        # Check username
                        if search_text in emp.get("username", "").lower():
                            filtered_employees.append(emp)
                            continue
                            
                        # Check ID
                        if search_text in str(emp.get("id", "")).lower():
                            filtered_employees.append(emp)
                            continue
                            
                        # Check role
                        if search_text in emp.get("role", "").lower():
                            filtered_employees.append(emp)
                            continue
                            
                        # Check branch name if available
                        branch_name = self.get_branch_name(emp.get("branch_id"))
                        if search_text in branch_name.lower():
                            filtered_employees.append(emp)
                            continue
                    
                    employees = filtered_employees

                # تعبئة الجدول بالبيانات المصفاة
                self.employees_table.setRowCount(len(employees))
                for i, emp in enumerate(employees):
                    username_item = QTableWidgetItem(emp.get("username", ""))
                    username_item.setData(Qt.ItemDataRole.UserRole, emp)  # Store employee data
                    self.employees_table.setItem(i, 0, username_item)
                    
                    # تحويل الدور إلى عربي
                    role_arabic = {
                        "employee": "موظف",
                        "branch_manager": "مدير فرع",
                        "director": "مدير النظام"
                    }.get(emp.get("role", ""), emp.get("role", ""))
                    self.employees_table.setItem(i, 1, QTableWidgetItem(role_arabic))
                    
                    # الحصول على اسم الفرع من الـ branch_id
                    branch_name = self.get_branch_name(emp.get("branch_id"))
                    self.employees_table.setItem(i, 2, QTableWidgetItem(branch_name))
                    
                    # تاريخ الإنشاء
                    self.employees_table.setItem(i, 3, QTableWidgetItem(emp.get("created_at", "")))
                    
                    # الحالة
                    status_item = QTableWidgetItem("نشط")
                    status_item.setForeground(QColor("#27ae60"))
                    self.employees_table.setItem(i, 4, status_item)

                # عرض رسالة إذا لم توجد نتائج
                if not employees:
                    self.show_no_results_message()
                    
            else:
                QMessageBox.warning(self, "خطأ", f"فشل تحميل البيانات: {response.status_code}")

        except Exception as e:
            print(f"Error in filter_employees: {e}")
            QMessageBox.warning(self, "خطأ", f"خطأ في التصفية: {str(e)}")

    def show_no_results_message(self):
        self.employees_table.setRowCount(1)
        no_results = QTableWidgetItem("لا توجد نتائج")
        no_results.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self.employees_table.setItem(0, 0, no_results)
        self.employees_table.setSpan(0, 0, 1, 5)
            
            
    def get_branch_name(self, branch_id):
        if not branch_id:
            return "غير محدد"
        
        # البحث في القائمة المنسدلة للفروع
        for index in range(self.branch_filter.count()):
            if self.branch_filter.itemData(index) == branch_id:
                return self.branch_filter.itemText(index)
        
        # إذا لم يتم العثور، جلب الاسم من الخادم
        try:
            headers = {"Authorization": f"Bearer {self.token}"} if self.token else {}
            response = requests.get(f"{self.api_url}/branches/{branch_id}", headers=headers)
            if response.status_code == 200:
                return response.json().get("name", "غير محدد")
        except:
            pass
        
        return "غير محدد"    
    
    def filter_transactions(self):
        """Filter transactions by branch, type, and status."""
        # Get branch ID from the branch filter dropdown
        branch_id = self.transaction_branch_filter.currentData()
        # Ensure branch_id is not the API URL and is valid
        if branch_id == self.api_url or branch_id == "":
            branch_id = None
            
        # Map filter type from dropdown index to API parameter
        filter_type_map = {
            0: "all",
            1: "incoming",
            2: "outgoing",
            3: "branch_related"
        }
        
        filter_type = filter_type_map.get(self.transaction_type_filter.currentIndex(), "all")
        
        # Get status filter value
        status = self.status_filter.currentData()
        # Only pass status if it's not "all"
        if status == "all":
            status = None
        
        # Reset pagination to first page when applying filters
        self.current_page_transactions = 1
        
        # Load transactions with all filter parameters
        self.load_transactions(branch_id, filter_type, status)
    
    def add_branch(self):
        """Open dialog to add a new branch."""
        dialog = AddBranchDialog(self.token, self)
        if dialog.exec() == 1:  # If dialog was accepted
            self.load_branches()  # Refresh the branches list
            self.load_branches_for_filter()  # Refresh branch filters
            self.load_dashboard_data()  # Refresh dashboard data
    
    def edit_branch(self):
        """Open dialog to edit the selected branch."""
        selected_rows = self.branches_table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "تنبيه", "الرجاء تحديد فرع للتعديل")
            return
        
        row = selected_rows[0].row()
        branch_data = self.branches_table.item(row, 0).data(Qt.ItemDataRole.UserRole)
        
        # Open edit dialog
        from ui.branch_management_improved import EditBranchDialog
        dialog = EditBranchDialog(branch_data, self.token, self)
        if dialog.exec() == 1:  # If dialog was accepted
            self.load_branches()  # Refresh the branches list
            self.load_branches_for_filter()  # Refresh branch filters
    
    def delete_branch(self):
        """Delete the selected branch."""
        selected_rows = self.branches_table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "تنبيه", "الرجاء تحديد فرع للحذف")
            return
        
        row = selected_rows[0].row()
        branch_data = self.branches_table.item(row, 0).data(Qt.ItemDataRole.UserRole)
        
        # Ensure we have valid branch data
        if not branch_data or not isinstance(branch_data, dict):
            QMessageBox.warning(self, "خطأ", "بيانات الفرع غير صالحة")
            return
        
        branch_id = branch_data.get("id")
        branch_name = branch_data.get("name", "هذا الفرع")
        
        # Confirm deletion
        reply = QMessageBox.question(
            self, 
            "تأكيد الحذف", 
            f"هل أنت متأكد من حذف الفرع '{branch_name}'؟",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            try:
                headers = {
                    "Authorization": f"Bearer {self.token}",
                    "Content-Type": "application/json"
                } if self.token else {}
                
                response = requests.delete(
                    f"{self.api_url}/branches/{branch_id}/", 
                    headers=headers
                )
                
                if response.status_code == 204:
                    QMessageBox.information(self, "نجاح", "تم حذف الفرع بنجاح")
                    self.load_branches()  # Refresh the branches list
                    self.load_branches_for_filter()  # Refresh branch filters
                    self.load_dashboard_data()  # Refresh dashboard data
                else:
                    # Try to get detailed error message
                    error_msg = f"فشل حذف الفرع: رمز الحالة {response.status_code}"
                    try:
                        error_data = response.json()
                        if "detail" in error_data:
                            error_msg = error_data["detail"]
                        elif "message" in error_data:
                            error_msg = error_data["message"]
                    except:
                        pass
                    
                    QMessageBox.warning(self, "خطأ", error_msg)
                    
            except requests.exceptions.RequestException as e:
                QMessageBox.warning(
                    self, 
                    "خطأ في الاتصال", 
                    f"تعذر الاتصال بالخادم: {str(e)}"
                )
            except Exception as e:
                QMessageBox.warning(
                    self, 
                    "خطأ غير متوقع", 
                    f"حدث خطأ غير متوقع: {str(e)}"
                )
    
    def view_branch(self):
        """View details of the selected branch."""
        selected_rows = self.branches_table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "تنبيه", "الرجاء تحديد فرع للعرض")
            return
        
        row = selected_rows[0].row()
        branch_data = self.branches_table.item(row, 0).data(Qt.ItemDataRole.UserRole)
        
        # Create a dialog to display branch details
        dialog = QDialog(self)
        dialog.setWindowTitle(f"تفاصيل الفرع: {branch_data.get('name', '')}")
        dialog.setGeometry(150, 150, 600, 400)
        dialog.setStyleSheet("""
            QDialog {
                background-color: #f5f5f5;
                font-family: Arial;
            }
            QLabel {
                color: #333;
            }
        """)
        
        layout = QVBoxLayout()
        
        # Branch details
        details_group = ModernGroupBox("معلومات الفرع", "#3498db")
        details_layout = QFormLayout()
        
        details_layout.addRow("رمز الفرع:", QLabel(branch_data.get("branch_id", "")))
        details_layout.addRow("اسم الفرع:", QLabel(branch_data.get("name", "")))
        details_layout.addRow("الموقع:", QLabel(branch_data.get("location", "")))
        details_layout.addRow("المحافظة:", QLabel(branch_data.get("governorate", "")))
        
        details_group.setLayout(details_layout)
        layout.addWidget(details_group)
        
        # Branch employees
        employees_group = ModernGroupBox("موظفي الفرع", "#2ecc71")
        employees_layout = QVBoxLayout()
        
        employees_table = QTableWidget()
        employees_table.setColumnCount(3)
        employees_table.setHorizontalHeaderLabels(["اسم المستخدم", "الدور", "تاريخ الإنشاء"])
        employees_table.horizontalHeader().setStretchLastSection(True)
        employees_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        
        try:
            headers = {"Authorization": f"Bearer {self.token}"} if self.token else {}
            response = requests.get(
                f"{self.api_url}/branches/{branch_data.get('id')}/employees/", 
                headers=headers
            )
            
            if response.status_code == 200:
                employees = response.json()
                employees_table.setRowCount(len(employees))
                
                for i, employee in enumerate(employees):
                    employees_table.setItem(i, 0, QTableWidgetItem(employee.get("username", "")))
                    
                    # Map role to Arabic
                    role = employee.get("role", "")
                    role_arabic = "موظف"
                    if role == "director":
                        role_arabic = "مدير"
                    elif role == "branch_manager":
                        role_arabic = "مدير فرع"
                    
                    employees_table.setItem(i, 1, QTableWidgetItem(role_arabic))
                    employees_table.setItem(i, 2, QTableWidgetItem(employee.get("created_at", "")))
        except Exception as e:
            print(f"Error loading branch employees: {e}")
        
        employees_layout.addWidget(employees_table)
        employees_group.setLayout(employees_layout)
        layout.addWidget(employees_group)
        
        # Branch statistics
        stats_group = ModernGroupBox("إحصائيات الفرع", "#e74c3c")
        stats_layout = QFormLayout()
        
        try:
            headers = {"Authorization": f"Bearer {self.token}"} if self.token else {}
            
            # Get employee stats
            emp_response = requests.get(
                f"{self.api_url}/branches/{branch_data.get('id')}/employees/stats/", 
                headers=headers
            )
            if emp_response.status_code == 200:
                emp_data = emp_response.json()
                stats_layout.addRow("عدد الموظفين:", QLabel(str(emp_data.get("total", 0))))
            
            # Get transaction stats
            trans_response = requests.get(
                f"{self.api_url}/branches/{branch_data.get('id')}/transactions/stats/", 
                headers=headers
            )
            if trans_response.status_code == 200:
                trans_data = trans_response.json()
                stats_layout.addRow("عدد التحويلات:", QLabel(str(trans_data.get("total", 0))))
                stats_layout.addRow("إجمالي المبالغ:", QLabel(f"{trans_data.get('total_amount', 0):,.2f}"))
        except Exception as e:
            print(f"Error loading branch statistics: {e}")
        
        stats_group.setLayout(stats_layout)
        layout.addWidget(stats_group)
        
        # Close button
        close_button = ModernButton("إغلاق", color="#e74c3c")
        close_button.clicked.connect(dialog.accept)
        layout.addWidget(close_button)
        
        dialog.setLayout(layout)
        dialog.exec()
    
    def add_employee(self):
        """Open dialog to add a new employee."""
        dialog = AddEmployeeDialog(
            is_admin=True,
            branch_id=None,
            token=self.token,
        )
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_employees()
            self.load_dashboard_data()
    
    def edit_employee(self):
        """Open dialog to edit the selected employee."""
        selected_rows = self.employees_table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "تنبيه", "الرجاء تحديد موظف للتعديل")
            return
        
        row = selected_rows[0].row()
        
        # التحقق من وجود العنصر في الجدول
        employee_item = self.employees_table.item(row, 0)
        if not employee_item:
            QMessageBox.warning(self, "خطأ", "الخليّة المحددة غير موجودة")
            return
        
        # استرجاع بيانات الموظف مع التحقق من وجودها
        employee_data = employee_item.data(Qt.ItemDataRole.UserRole)
        if not employee_data or not isinstance(employee_data, dict):
            QMessageBox.warning(self, "خطأ", "بيانات الموظف غير صالحة")
            return
        
        # التحقق من وجود الحقول الأساسية
        if "id" not in employee_data or "username" not in employee_data:
            QMessageBox.warning(self, "خطأ", "بيانات الموظف ناقصة أو غير صحيحة")
            return
        
        # فتح نافذة التعديل
        from ui.user_management_improved import EditEmployeeDialog
        dialog = EditEmployeeDialog(employee_data, self.token, self)
        if dialog.exec() == 1:  # If dialog was accepted
            self.load_employees()  # Refresh the employees list
    
    def delete_employee(self):
        """Delete the selected employee."""
        selected_rows = self.employees_table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "تنبيه", "الرجاء تحديد موظف للحذف")
            return
        
        row = selected_rows[0].row()
        
        # التحقق من وجود العنصر في الجدول
        employee_item = self.employees_table.item(row, 0)
        if not employee_item:
            QMessageBox.warning(self, "خطأ", "الخليّة المحددة غير موجودة")
            return
        
        # استرجاع بيانات الموظف مع التحقق من وجودها
        employee_data = employee_item.data(Qt.ItemDataRole.UserRole)
        if not employee_data or not isinstance(employee_data, dict):
            QMessageBox.warning(self, "خطأ", "بيانات الموظف غير صالحة")
            return
        
        employee_id = employee_data.get("id")
        employee_name = employee_data.get("username", "")
        
        # تأكيد الحذف
        reply = QMessageBox.question(
            self, 
            "تأكيد الحذف", 
            f"هل أنت متأكد من حذف الموظف '{employee_name}'؟",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            try:
                headers = {"Authorization": f"Bearer {self.token}"} if self.token else {}
                response = requests.delete(f"{self.api_url}/users/{employee_id}", headers=headers)
                
                if response.status_code in [200, 204]:
                    QMessageBox.information(self, "نجاح", "تم حذف الموظف بنجاح")
                    self.load_employees()
                    self.load_dashboard_data()
                else:
                    error_msg = f"فشل حذف الموظف: رمز الحالة {response.status_code}"
                    try:
                        error_data = response.json()
                        error_msg = error_data.get("detail", error_msg)
                    except:
                        pass
                    QMessageBox.warning(self, "خطأ", error_msg)
            except Exception as e:
                QMessageBox.warning(self, "خطأ", f"تعذر حذف الموظف: {str(e)}")
    
    def search_user(self):
        """Open user search dialog."""
        dialog = UserSearchDialog(self.token, self)
        dialog.exec()
    
    def reset_password(self):
        """Reset password for the selected employee."""
        selected_rows = self.employees_table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "تنبيه", "الرجاء تحديد موظف لإعادة تعيين كلمة المرور")
            return
        
        row = selected_rows[0].row()
        employee_data = self.employees_table.item(row, 0).data(Qt.ItemDataRole.UserRole)
        employee_username = employee_data.get("username", "")
        
        # Create a dialog to reset password
        dialog = QDialog(self)
        dialog.setWindowTitle(f"إعادة تعيين كلمة المرور: {employee_username}")
        dialog.setGeometry(150, 150, 400, 200)
        dialog.setStyleSheet("""
            QDialog {
                background-color: #f5f5f5;
                font-family: Arial;
            }
            QLabel {
                color: #333;
            }
            QLineEdit {
                border: 1px solid #ccc;
                border-radius: 5px;
                padding: 5px;
                background-color: white;
            }
        """)
        
        layout = QVBoxLayout()
        
        # Password fields
        form_layout = QFormLayout()
        
        new_password_label = QLabel("كلمة المرور الجديدة:")
        new_password_input = QLineEdit()
        new_password_input.setEchoMode(QLineEdit.EchoMode.Password)
        form_layout.addRow(new_password_label, new_password_input)
        
        confirm_password_label = QLabel("تأكيد كلمة المرور:")
        confirm_password_input = QLineEdit()
        confirm_password_input.setEchoMode(QLineEdit.EchoMode.Password)
        form_layout.addRow(confirm_password_label, confirm_password_input)
        
        layout.addLayout(form_layout)
        
        # Buttons
        buttons_layout = QHBoxLayout()
        
        cancel_button = ModernButton("إلغاء", color="#e74c3c")
        cancel_button.clicked.connect(dialog.reject)
        buttons_layout.addWidget(cancel_button)
        
        save_button = ModernButton("حفظ", color="#2ecc71")
        
        def reset_password_action():
            new_password = new_password_input.text()
            confirm_password = confirm_password_input.text()
            
            if not new_password:
                QMessageBox.warning(dialog, "تنبيه", "الرجاء إدخال كلمة المرور الجديدة")
                return
            
            if new_password != confirm_password:
                QMessageBox.warning(dialog, "تنبيه", "كلمة المرور وتأكيدها غير متطابقين")
                return
            
            try:
                headers = {"Authorization": f"Bearer {self.token}"} if self.token else {}
                data = {
                    "username": employee_username,
                    "new_password": new_password
                }
                response = requests.post(f"{self.api_url}/reset-password/", json=data, headers=headers)
                
                if response.status_code == 200:
                    QMessageBox.information(dialog, "نجاح", "تم إعادة تعيين كلمة المرور بنجاح")
                    dialog.accept()
                else:
                    error_msg = f"فشل إعادة تعيين كلمة المرور: رمز الحالة {response.status_code}"
                    try:
                        error_data = response.json()
                        if "detail" in error_data:
                            error_msg = error_data["detail"]
                    except:
                        pass
                    
                    QMessageBox.warning(dialog, "خطأ", error_msg)
            except Exception as e:
                print(f"Error resetting password: {e}")
                QMessageBox.warning(dialog, "خطأ", f"تعذر إعادة تعيين كلمة المرور: {str(e)}")
        
        save_button.clicked.connect(reset_password_action)
        buttons_layout.addWidget(save_button)
        
        layout.addLayout(buttons_layout)
        
        dialog.setLayout(layout)
        dialog.exec()
    
    def view_transaction(self):
        """View details of the selected transaction."""
        selected_rows = self.transactions_table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "تنبيه", "الرجاء تحديد تحويل للعرض")
            return
        
        row = selected_rows[0].row()
        transaction = self.transactions_table.item(row, 0).data(Qt.ItemDataRole.UserRole)
        
        # Create a dialog to display transaction details
        dialog = QDialog(self)
        dialog.setWindowTitle("تفاصيل التحويل")
        dialog.setGeometry(150, 150, 500, 400)
        dialog.setStyleSheet("""
            QDialog {
                background-color: #f5f5f5;
                font-family: Arial;
            }
            QLabel {
                color: #333;
            }
        """)
        
        layout = QVBoxLayout()
        
        # Title
        title = QLabel("تفاصيل التحويل")
        title.setFont(QFont("Arial", 16, QFont.Weight.Bold))
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setStyleSheet("color: #2c3e50; margin-bottom: 20px;")
        layout.addWidget(title)
        
        # Transaction details
        details_group = ModernGroupBox("معلومات التحويل", "#3498db")
        details_layout = QVBoxLayout()
        
        # Format transaction details
        details_text = f"""
        <b>رقم التحويل:</b> {transaction.get('id', '')}<br>
        <b>التاريخ:</b> {transaction.get('date', '')}<br>
        <br>
        <b>المرسل:</b> {transaction.get('sender', '')}<br>
        <b>رقم هاتف المرسل:</b> {transaction.get('sender_mobile', '')}<br>
        <b>محافظة المرسل:</b> {transaction.get('sender_governorate', '')}<br>
        <b>موقع المرسل:</b> {transaction.get('sender_location', '')}<br>
        <br>
        <b>المستلم:</b> {transaction.get('receiver', '')}<br>
        <b>رقم هاتف المستلم:</b> {transaction.get('receiver_mobile', '')}<br>
        <b>محافظة المستلم:</b> {transaction.get('receiver_governorate', '')}<br>
        <b>موقع المستلم:</b> {transaction.get('receiver_location', '')}<br>
        <br>
        <b>المبلغ:</b> {transaction.get('amount', '')} {transaction.get('currency', '')}<br>
        <b>الرسالة:</b> {transaction.get('message', '')}<br>
        <br>
        <b>الفرع:</b> {transaction.get('branch_governorate', '')}<br>
        <b>الموظف:</b> {transaction.get('employee_name', '')}<br>
        <b>الحالة:</b> {transaction.get('status', '')}
        """
        
        details_label = QLabel(details_text)
        details_label.setTextFormat(Qt.TextFormat.RichText)
        details_label.setWordWrap(True)
        details_layout.addWidget(details_label)
        
        details_group.setLayout(details_layout)
        layout.addWidget(details_group)
        
        # Close button
        close_button = ModernButton("إغلاق", color="#e74c3c")
        close_button.clicked.connect(dialog.accept)
        layout.addWidget(close_button)
        
        dialog.setLayout(layout)
        dialog.exec()
    
    def update_transaction_status(self):
        """Update status of the selected transaction."""
        selected_rows = self.transactions_table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "تنبيه", "الرجاء تحديد تحويل لتحديث الحالة")
            return
        
        row = selected_rows[0].row()
        transaction = self.transactions_table.item(row, 0).data(Qt.ItemDataRole.UserRole)
        transaction_id = transaction.get('id', '')
        
        # Create a dialog to update status
        dialog = QDialog(self)
        dialog.setWindowTitle("تحديث حالة التحويل")
        dialog.setGeometry(150, 150, 400, 200)
        dialog.setStyleSheet("""
            QDialog {
                background-color: #f5f5f5;
                font-family: Arial;
            }
            QLabel {
                color: #333;
            }
            QComboBox {
                border: 1px solid #ccc;
                border-radius: 5px;
                padding: 5px;
                background-color: white;
            }
        """)
        
        layout = QVBoxLayout()
        
        # Status selection
        form_layout = QFormLayout()
        
        status_label = QLabel("الحالة:")
        status_combo = QComboBox()
        status_combo.addItems(["قيد الانتظار", "تم الاستلام", "ملغي"])
        
        # Set current status
        current_status = transaction.get('status', '')
        if current_status == "completed":
            status_combo.setCurrentText("تم الاستلام")
        elif current_status == "cancelled":
            status_combo.setCurrentText("ملغي")
        else:
            status_combo.setCurrentText("قيد الانتظار")
        
        form_layout.addRow(status_label, status_combo)
        
        layout.addLayout(form_layout)
        
        # Buttons
        buttons_layout = QHBoxLayout()
        
        cancel_button = ModernButton("إلغاء", color="#e74c3c")
        cancel_button.clicked.connect(dialog.reject)
        buttons_layout.addWidget(cancel_button)
        
        save_button = ModernButton("حفظ", color="#2ecc71")
        
        def update_status_action():
            # Map Arabic status to English
            status_map = {
                "قيد الانتظار": "processing",
                "تم الاستلام": "completed",
                "ملغي": "cancelled"
            }
            
            new_status = status_map.get(status_combo.currentText(), "processing")
            
            try:
                headers = {"Authorization": f"Bearer {self.token}"} if self.token else {}
                data = {
                    "transaction_id": transaction_id,
                    "status": new_status
                }
                response = requests.post(f"{self.api_url}/update-transaction-status/", json=data, headers=headers)
                
                if response.status_code == 200:
                    QMessageBox.information(dialog, "نجاح", "تم تحديث حالة التحويل بنجاح")
                    dialog.accept()
                    self.load_transactions()  # Refresh the transactions list
                else:
                    error_msg = f"فشل تحديث حالة التحويل: رمز الحالة {response.status_code}"
                    try:
                        error_data = response.json()
                        if "detail" in error_data:
                            error_msg = error_data["detail"]
                    except:
                        pass
                    
                    QMessageBox.warning(dialog, "خطأ", error_msg)
            except Exception as e:
                print(f"Error updating transaction status: {e}")
                QMessageBox.warning(dialog, "خطأ", f"تعذر تحديث حالة التحويل: {str(e)}")
        
        save_button.clicked.connect(update_status_action)
        buttons_layout.addWidget(save_button)
        
        layout.addLayout(buttons_layout)
        
        dialog.setLayout(layout)
        dialog.exec()
    
    def print_receipt(self):
        """Print receipt for the selected transaction."""
        selected_rows = self.transactions_table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "تنبيه", "الرجاء تحديد تحويل لطباعة الإيصال")
            return
        
        row = selected_rows[0].row()
        transaction = self.transactions_table.item(row, 0).data(Qt.ItemDataRole.UserRole)
        
        try:
            # Format transaction ID with dashes like in the image (1925-5814-5221)
            # If the ID is numeric, format it with dashes
            formatted_id = str(transaction.get('id', ''))
            if formatted_id.isdigit() and len(formatted_id) >= 4:
                id_parts = []
                id_str = formatted_id
                while id_str:
                    if len(id_str) > 4:
                        id_parts.append(id_str[:4])
                        id_str = id_str[4:]
                    else:
                        id_parts.append(id_str)
                        id_str = ""
                formatted_id = "-".join(id_parts)
            
            # Convert amount to Arabic words
            amount = transaction.get('amount', '0')
            currency = transaction.get('currency', 'ليرة سورية')
            amount_in_arabic = number_to_arabic_words(amount, currency)
            
            # Determine transaction type (incoming or outgoing)
            transaction_type = "تسليم" if transaction.get('type') == 'received' else "إرسال"
            transaction_type_full = "تحويل وارد" if transaction.get('type') == 'received' else "تحويل صادر"
            
            # Create print dialog
            print_dialog = QDialog(self)
            print_dialog.setWindowTitle("طباعة التحويل")
            print_dialog.setFixedSize(600, 700)
            
            layout = QVBoxLayout()
            
            # Create printable content with improved layout
            content = QTextEdit()
            content.setReadOnly(True)
            content.setHtml(f"""
                <div style='font-family: Arial; direction: rtl; background-color: #f9f9f9; padding: 0; margin: 0;'>
                    <!-- Header -->
                    <div style='text-align: right; padding: 10px; background-color: #e8f5e9; border-bottom: 1px solid #ddd;'>
                        <h2 style='margin: 0; color: #333;'>طباعة التحويل</h2>
                    </div>
                    
                    <!-- Transaction Number Row -->
                    <div style='margin: 15px 0; text-align: center;'>
                        <div style='display: flex; justify-content: space-between; align-items: center;'>
                            <div style='flex: 1; text-align: right;'>
                                <strong style='color: #333;'>الرقم</strong>
                            </div>
                            <div style='flex: 3; text-align: center;'>
                                <strong style='font-size: 16px;'>{formatted_id}</strong>
                            </div>
                            <div style='flex: 1; text-align: center; background-color: #e8f5e9; padding: 5px; border-radius: 4px;'>
                                <strong style='color: #2e7d32;'>{transaction_type}</strong>
                            </div>
                        </div>
                    </div>
                    
                    <!-- Main Content -->
                    <div style='margin: 0 10px;'>
                        <!-- From/To Row -->
                        <div style='display: flex; justify-content: space-between; margin-bottom: 8px; border-bottom: 1px solid #eee; padding-bottom: 8px;'>
                            <div style='flex: 1; text-align: right;'>
                                <strong style='color: #555;'>من</strong>
                            </div>
                            <div style='flex: 5; text-align: left;'>
                                {self.branch_id_to_name.get(transaction.get('branch_id'), '')}
                            </div>
                        </div>
                        
                        <div style='display: flex; justify-content: space-between; margin-bottom: 8px; border-bottom: 1px solid #eee; padding-bottom: 8px;'>
                            <div style='flex: 1; text-align: right;'>
                                <strong style='color: #555;'>إلى</strong>
                            </div>
                            <div style='flex: 5; text-align: left;'>
                                {self.branch_id_to_name.get(transaction.get('destination_branch_id'), '')}
                            </div>
                        </div>
                        
                        <!-- Sender Row -->
                        <div style='display: flex; justify-content: space-between; margin-bottom: 8px; border-bottom: 1px solid #eee; padding-bottom: 8px;'>
                            <div style='flex: 1; text-align: right;'>
                                <strong style='color: #555;'>المرسل</strong>
                            </div>
                            <div style='flex: 5; text-align: left;'>
                                {transaction.get('sender_name', '')}
                            </div>
                        </div>
                        
                        <!-- Receiver Row -->
                        <div style='display: flex; justify-content: space-between; margin-bottom: 8px; border-bottom: 1px solid #eee; padding-bottom: 8px;'>
                            <div style='flex: 1; text-align: right;'>
                                <strong style='color: #555;'>المفوض</strong>
                            </div>
                            <div style='flex: 5; text-align: left;'>
                                {transaction.get('receiver_name', '')}
                            </div>
                        </div>
                        
                        <!-- Contact Row -->
                        <div style='display: flex; justify-content: space-between; margin-bottom: 8px; border-bottom: 1px solid #eee; padding-bottom: 8px;'>
                            <div style='flex: 1; text-align: right;'>
                                <strong style='color: #555;'>اتصال</strong>
                            </div>
                            <div style='flex: 5; text-align: left;'>
                                {transaction.get('receiver_mobile', '')}
                            </div>
                        </div>
                        
                        <!-- Beneficiary Row -->
                        <div style='display: flex; justify-content: space-between; margin-bottom: 8px; border-bottom: 1px solid #eee; padding-bottom: 8px;'>
                            <div style='flex: 1; text-align: right;'>
                                <strong style='color: #555;'>المستفيد</strong>
                            </div>
                            <div style='flex: 5; text-align: left;'>
                                {transaction.get('receiver_name', '')}
                            </div>
                        </div>
                        
                        <!-- Recipient Row -->
                        <div style='display: flex; justify-content: space-between; margin-bottom: 8px; border-bottom: 1px solid #eee; padding-bottom: 8px;'>
                            <div style='flex: 1; text-align: right;'>
                                <strong style='color: #555;'>المستلم</strong>
                            </div>
                            <div style='flex: 5; text-align: left;'>
                                {transaction.get('receiver_name', '')}
                            </div>
                        </div>
                        
                        <!-- Contact (Phone) Row -->
                        <div style='display: flex; justify-content: space-between; margin-bottom: 8px; border-bottom: 1px solid #eee; padding-bottom: 8px;'>
                            <div style='flex: 1; text-align: right;'>
                                <strong style='color: #555;'>اتصال</strong>
                            </div>
                            <div style='flex: 5; text-align: left;'>
                                {transaction.get('receiver_mobile', '')}
                            </div>
                        </div>
                        
                        <!-- Date Row -->
                        <div style='display: flex; justify-content: space-between; margin-bottom: 8px; border-bottom: 1px solid #eee; padding-bottom: 8px;'>
                            <div style='flex: 1; text-align: right;'>
                                <strong style='color: #555;'>التاريخ</strong>
                            </div>
                            <div style='flex: 5; text-align: left;'>
                                {transaction.get('date', '')}
                            </div>
                        </div>
                        
                        <!-- Amount Row -->
                        <div style='display: flex; justify-content: space-between; margin-bottom: 8px; border-bottom: 1px solid #eee; padding-bottom: 8px;'>
                            <div style='flex: 1; text-align: right;'>
                                <strong style='color: #555;'>المبلغ</strong>
                            </div>
                            <div style='flex: 5; text-align: left;'>
                                {transaction.get('amount', '')} {transaction.get('currency', '')}
                            </div>
                        </div>
                        
                        <!-- Amount in Words Row -->
                        <div style='margin-bottom: 15px; padding: 8px; background-color: #f5f5f5; border-radius: 4px;'>
                            <p style='margin: 0; text-align: right; color: #333;'>
                                {amount_in_arabic}
                            </p>
                        </div>
                    </div>
                    
                    <!-- Footer Note -->
                    <div style='margin: 15px 0; text-align: center; color: #4CAF50; font-weight: bold;'>
                        <p>يطلب تبليغ المستفيد برقم الحوالة</p>
                    </div>
                    
                    <!-- System Section -->
                    <hr style='border: none; border-top: 1px solid #ddd; margin: 20px 0;'>
                    
                    <div style='text-align: right; margin: 0 10px;'>
                        <h3 style='color: #4CAF50; margin-bottom: 15px;'>نسخة النظام</h3>
                        
                        <div style='margin-bottom: 8px;'>
                            <strong style='color: #555;'>بيانات فرع المرسل:</strong>
                            <span style='margin-right: 10px;'>{self.branch_id_to_name.get(transaction.get('branch_id'), '')}</span>
                        </div>
                        
                        <div style='margin-bottom: 8px;'>
                            <strong style='color: #555;'>بيانات فرع المستلم:</strong>
                            <span style='margin-right: 10px;'>{self.branch_id_to_name.get(transaction.get('destination_branch_id'), '')}</span>
                        </div>
                        
                        <div style='margin-bottom: 8px;'>
                            <strong style='color: #555;'>اسم الموظف:</strong>
                            <span style='margin-right: 10px;'>{transaction.get('employee_name', '')}</span>
                        </div>
                        
                        <div style='margin-bottom: 8px;'>
                            <strong style='color: #555;'>نوع التحويل:</strong>
                            <span style='margin-right: 10px;'>{transaction_type_full}</span>
                        </div>
                        
                        <div style='margin-bottom: 8px;'>
                            <strong style='color: #555;'>حالة التحويل:</strong>
                            <span style='margin-right: 10px;'>{transaction.get('status', '')}</span>
                        </div>
                        
                        <div style='margin-top: 30px; border-top: 1px dashed #ddd; padding-top: 15px;'>
                            <p><strong>اسم العميل الكامل:</strong> ________________________</p>
                            <p><strong>التوقيع:</strong> ________________________</p>
                        </div>
                    </div>
                </div>
            """)
            
            # Print button
            print_btn = QPushButton("طباعة")
            print_btn.setStyleSheet("""
                QPushButton {
                    background-color: #4CAF50;
                    color: white;
                    border: none;
                    padding: 8px 16px;
                    font-size: 14px;
                    border-radius: 4px;
                }
                QPushButton:hover {
                    background-color: #45a049;
                }
            """)
            print_btn.clicked.connect(lambda: self.send_to_printer(content))
            
            layout.addWidget(content)
            layout.addWidget(print_btn)
            print_dialog.setLayout(layout)
            print_dialog.exec()
            
        except Exception as e:
            print(f"Error printing receipt: {e}")
            QMessageBox.warning(self, "خطأ", f"حدث خطأ أثناء طباعة الإيصال: {str(e)}")
    
    def send_to_printer(self, content):
        """Handle actual printing functionality."""
        from PyQt6.QtPrintSupport import QPrinter, QPrintDialog
        
        printer = QPrinter()
        print_dialog = QPrintDialog(printer, self)
        
        if print_dialog.exec() == QDialog.DialogCode.Accepted:
            content.print(printer)
    
    def search_transaction(self):
        """Open transaction search dialog."""
        from ui.user_search import UserSearchDialog
        dialog = UserSearchDialog(token=self.token, parent=self, received=True)
        dialog.exec()
    
    def generate_report(self):
        """Generate a report based on the selected options."""
        report_type_map = {
            "تقرير التحويلات": "transactions",
            "تقرير الفروع": "branches",
            "تقرير الموظفين": "employees"
        }
        
        report_type = report_type_map.get(self.report_type.currentText(), "transactions")
        from_date = self.from_date.date().toString("yyyy-MM-dd")
        to_date = self.to_date.date().toString("yyyy-MM-dd")
        branch_id = self.report_branch_filter.currentData()
        
        # Ensure branch_id is not the API URL
        if branch_id == self.api_url:
            branch_id = None
            
        try:
            headers = {"Authorization": f"Bearer {self.token}"} if self.token else {}
            params = {
                "date_from": from_date,
                "date_to": to_date
            }
            
            if branch_id:
                params["branch_id"] = branch_id
            
            # For transactions, we'll use the regular transactions endpoint to get more detailed data
            if report_type == "transactions":
                response = requests.get(
                    f"{self.api_url}/transactions/", 
                    headers=headers,
                    params=params
                )
            else:
                response = requests.get(
                    f"{self.api_url}/reports/{report_type}/", 
                    headers=headers,
                    params=params
                )
            
            if response.status_code == 200:
                data = response.json()
                
                # Get items based on report type
                if report_type == "transactions":
                    items = data.get("transactions", [])
                else:
                    items = data.get("items", [])
                
                # Set up table columns based on report type
                if report_type == "transactions":
                    # Enhanced columns for transactions as requested by user
                    self.report_table.setColumnCount(10)
                    self.report_table.setHorizontalHeaderLabels([
                        "النوع", "رقم التحويل", "المرسل", "المستلم", "المبلغ", 
                        "التاريخ", "الحالة", "الفرع المرسل", "الفرع المستلم", "اسم الموظف"
                    ])
                elif report_type == "branches":
                    self.report_table.setColumnCount(5)
                    self.report_table.setHorizontalHeaderLabels([
                        "رمز الفرع", "اسم الفرع", "الموقع", "المحافظة", "الحالة"
                    ])
                elif report_type == "employees":
                    self.report_table.setColumnCount(4)
                    self.report_table.setHorizontalHeaderLabels([
                        "اسم المستخدم", "الدور", "الفرع", "تاريخ الإنشاء"
                    ])
                
                # Fill table with data
                self.report_table.setRowCount(len(items))
                
                # Load branch names if not cached
                if not hasattr(self, 'branch_id_to_name') or not self.branch_id_to_name:
                    self.branch_id_to_name = {}
                    branches_response = requests.get(f"{self.api_url}/branches/", headers=headers)
                    if branches_response.status_code == 200:
                        branches = branches_response.json().get("branches", [])
                        self.branch_id_to_name = {b["id"]: b["name"] for b in branches}
                
                for i, item in enumerate(items):
                    if report_type == "transactions":
                        # Transaction Type
                        transaction_type = self.determine_transaction_type(item)
                        type_item = QTableWidgetItem(transaction_type)
                        if transaction_type == "داخلي":
                            type_item.setForeground(QColor(0, 128, 0))  # Green for internal
                        elif transaction_type == "صادر":
                            type_item.setForeground(QColor(255, 0, 0))  # Red for outgoing
                        elif transaction_type == "وارد":
                            type_item.setForeground(QColor(0, 0, 255))  # Blue for incoming
                        self.report_table.setItem(i, 0, type_item)
                        
                        # Transaction ID
                        trans_id = str(item.get("id", ""))
                        id_item = QTableWidgetItem(trans_id[:8] + "..." if len(trans_id) > 8 else trans_id)
                        id_item.setToolTip(trans_id)
                        self.report_table.setItem(i, 1, id_item)
                        
                        # Sender/Receiver
                        self.report_table.setItem(i, 2, QTableWidgetItem(item.get("sender", "")))
                        self.report_table.setItem(i, 3, QTableWidgetItem(item.get("receiver", "")))
                        
                        # Amount
                        amount = item.get("amount", 0)
                        currency = item.get("currency", "ليرة سورية")
                        amount_item = QTableWidgetItem(f"{amount:,.2f} {currency}")
                        amount_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                        self.report_table.setItem(i, 4, amount_item)
                        
                        # Date
                        date_str = item.get("date", "")
                        self.report_table.setItem(i, 5, QTableWidgetItem(date_str))
                        
                        # Status
                        status = item.get("status", "").lower()
                        status_ar = self.get_status_arabic(status)
                        status_item = QTableWidgetItem(status_ar)
                        status_item.setBackground(self.get_status_color(status))
                        status_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                        self.report_table.setItem(i, 6, status_item)
                        
                        # Sending Branch
                        branch_id = item.get("branch_id")
                        sending_branch = self.branch_id_to_name.get(branch_id, f"الفرع {branch_id}" if branch_id else "غير معروف")
                        self.report_table.setItem(i, 7, QTableWidgetItem(sending_branch))
                        
                        # Receiving Branch
                        dest_branch_id = item.get("destination_branch_id")
                        receiving_branch = self.branch_id_to_name.get(dest_branch_id, f"الفرع {dest_branch_id}" if dest_branch_id else "غير معروف")
                        self.report_table.setItem(i, 8, QTableWidgetItem(receiving_branch))
                        
                        # Employee Name
                        self.report_table.setItem(i, 9, QTableWidgetItem(item.get("employee_name", "")))
                        
                        # Store transaction data for potential use
                        self.report_table.item(i, 0).setData(Qt.ItemDataRole.UserRole, item)
                        
                    elif report_type == "branches":
                        self.report_table.setItem(i, 0, QTableWidgetItem(item.get("branch_id", "")))
                        self.report_table.setItem(i, 1, QTableWidgetItem(item.get("name", "")))
                        self.report_table.setItem(i, 2, QTableWidgetItem(item.get("location", "")))
                        self.report_table.setItem(i, 3, QTableWidgetItem(item.get("governorate", "")))
                        self.report_table.setItem(i, 4, QTableWidgetItem(item.get("status", "")))
                    elif report_type == "employees":
                        self.report_table.setItem(i, 0, QTableWidgetItem(item.get("username", "")))
                        self.report_table.setItem(i, 1, QTableWidgetItem(item.get("role", "")))
                        self.report_table.setItem(i, 2, QTableWidgetItem(str(item.get("branch_id", ""))))
                        self.report_table.setItem(i, 3, QTableWidgetItem(item.get("created_at", "")))
                
                # Adjust table appearance
                self.report_table.horizontalHeader().setStretchLastSection(True)
                self.report_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
                
                # Enable sorting
                self.report_table.setSortingEnabled(True)
                
                QMessageBox.information(self, "نجاح", "تم إنشاء التقرير بنجاح")
            else:
                QMessageBox.warning(self, "خطأ", f"فشل إنشاء التقرير: رمز الحالة {response.status_code}")
        except Exception as e:
            print(f"Error generating report: {e}")
            QMessageBox.warning(self, "خطأ", f"تعذر إنشاء التقرير: {str(e)}")
            
    def determine_transaction_type(self, transaction):
        """Determine transaction type based on branch information."""
        branch_id = transaction.get("branch_id")
        dest_branch_id = transaction.get("destination_branch_id")
        
        if branch_id and dest_branch_id:
            return "داخلي"  # Internal transfer between branches
        elif branch_id and not dest_branch_id:
            return "صادر"   # Outgoing transfer from branch
        elif not branch_id and dest_branch_id:
            return "وارد"   # Incoming transfer to branch
        else:
            return "غير معروف"  # Unknown type
    
    def export_pdf(self):
        """Export the current report as PDF."""
        if self.report_table.rowCount() == 0:
            QMessageBox.warning(self, "تنبيه", "لا توجد بيانات للتصدير. الرجاء إنشاء تقرير أولاً.")
            return
            
        try:
            from PyQt6.QtPrintSupport import QPrinter
            from PyQt6.QtGui import QTextDocument
            
            # Get file name from user
            file_path, _ = QFileDialog.getSaveFileName(
                self, "حفظ التقرير كـ PDF", "", "ملفات PDF (*.pdf)"
            )
            
            if not file_path:
                return  # User canceled
                
            # Add .pdf extension if not present
            if not file_path.lower().endswith('.pdf'):
                file_path += '.pdf'
                
            # Create printer
            printer = QPrinter(QPrinter.PrinterMode.HighResolution)
            printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
            printer.setOutputFileName(file_path)
            
            # Create HTML content
            html = "<html><head><style>"
            html += "table { width: 100%; border-collapse: collapse; direction: rtl; }"
            html += "th, td { border: 1px solid #000; padding: 8px; text-align: right; }"
            html += "th { background-color: #f2f2f2; }"
            html += "h1 { text-align: center; color: #2c3e50; }"
            html += ".status-completed { background-color: #d5f5e3; }"  # Light green
            html += ".status-processing { background-color: #d6eaf8; }" # Light blue
            html += ".status-cancelled { background-color: #f5b7b1; }"  # Light red
            html += ".status-rejected { background-color: #f1948a; }"   # Darker red
            html += ".status-on_hold { background-color: #fdebd0; }"    # Light orange
            html += "</style></head><body>"
            
            # Add title based on report type
            report_title = self.report_type.currentText()
            html += f"<h1>{report_title}</h1>"
            
            # Add date range
            from_date = self.from_date.date().toString("yyyy-MM-dd")
            to_date = self.to_date.date().toString("yyyy-MM-dd")
            html += f"<p style='text-align: center;'>الفترة من {from_date} إلى {to_date}</p>"
            
            # Create table
            html += "<table><tr>"
            
            # Add headers
            for col in range(self.report_table.columnCount()):
                header_text = self.report_table.horizontalHeaderItem(col).text()
                html += f"<th>{header_text}</th>"
            html += "</tr>"
            
            # Add data rows
            for row in range(self.report_table.rowCount()):
                html += "<tr>"
                
                # Get status for styling
                status_item = None
                status_class = ""
                if self.report_type.currentText() == "تقرير التحويلات":
                    status_item = self.report_table.item(row, 6)  # Status column
                    if status_item:
                        status_text = status_item.text().lower()
                        if "مكتمل" in status_text:
                            status_class = "status-completed"
                        elif "قيد المعالجة" in status_text:
                            status_class = "status-processing"
                        elif "ملغي" in status_text:
                            status_class = "status-cancelled"
                        elif "مرفوض" in status_text:
                            status_class = "status-rejected"
                        elif "معلق" in status_text:
                            status_class = "status-on_hold"
                
                for col in range(self.report_table.columnCount()):
                    item = self.report_table.item(row, col)
                    text = item.text() if item else ""
                    
                    # Apply status class to the entire row for transaction reports
                    if self.report_type.currentText() == "تقرير التحويلات" and status_class:
                        html += f"<td class='{status_class}'>{text}</td>"
                    else:
                        html += f"<td>{text}</td>"
                html += "</tr>"
            
            html += "</table></body></html>"
            
            # Print to PDF
            document = QTextDocument()
            document.setHtml(html)
            document.print(printer)
            
            QMessageBox.information(self, "نجاح", f"تم تصدير التقرير بنجاح إلى:\n{file_path}")
            
        except Exception as e:
            print(f"Error exporting to PDF: {e}")
            QMessageBox.warning(self, "خطأ", f"تعذر تصدير التقرير: {str(e)}")
    
    def export_excel(self):
        """Export the current report as Excel."""
        if self.report_table.rowCount() == 0:
            QMessageBox.warning(self, "تنبيه", "لا توجد بيانات للتصدير. الرجاء إنشاء تقرير أولاً.")
            return
            
        try:
            # Get file name from user
            file_path, _ = QFileDialog.getSaveFileName(
                self, "حفظ التقرير كـ Excel", "", "ملفات Excel (*.xlsx)"
            )
            
            if not file_path:
                return  # User canceled
                
            # Add .xlsx extension if not present
            if not file_path.lower().endswith('.xlsx'):
                file_path += '.xlsx'
                
            # Create Excel workbook
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Set title based on report type
            report_title = self.report_type.currentText()
            ws.title = report_title
            
            # Add title
            ws.merge_cells('A1:G1')
            title_cell = ws['A1']
            title_cell.value = report_title
            title_cell.font = Font(size=16, bold=True)
            title_cell.alignment = Alignment(horizontal='center')
            
            # Add date range
            from_date = self.from_date.date().toString("yyyy-MM-dd")
            to_date = self.to_date.date().toString("yyyy-MM-dd")
            ws.merge_cells('A2:G2')
            date_cell = ws['A2']
            date_cell.value = f"الفترة من {from_date} إلى {to_date}"
            date_cell.alignment = Alignment(horizontal='center')
            
            # Add headers
            header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            header_font = Font(bold=True)
            
            for col in range(self.report_table.columnCount()):
                header_text = self.report_table.horizontalHeaderItem(col).text()
                cell = ws.cell(row=4, column=col+1, value=header_text)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='right')
            
            # Add data rows
            for row in range(self.report_table.rowCount()):
                for col in range(self.report_table.columnCount()):
                    item = self.report_table.item(row, col)
                    text = item.text() if item else ""
                    cell = ws.cell(row=row+5, column=col+1, value=text)
                    cell.alignment = Alignment(horizontal='right')
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            
            # Save the workbook
            wb.save(file_path)
            
            QMessageBox.information(self, "نجاح", f"تم تصدير التقرير بنجاح إلى:\n{file_path}")
            
        except ImportError:
            QMessageBox.warning(self, "خطأ", "مكتبة openpyxl غير متوفرة. الرجاء تثبيتها باستخدام pip install openpyxl")
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            QMessageBox.warning(self, "خطأ", f"تعذر تصدير التقرير: {str(e)}")
    
    def print_report(self):
        """Print the current report."""
        if self.report_table.rowCount() == 0:
            QMessageBox.warning(self, "تنبيه", "لا توجد بيانات للطباعة. الرجاء إنشاء تقرير أولاً.")
            return
            
        try:
            from PyQt6.QtPrintSupport import QPrinter, QPrintDialog
            from PyQt6.QtGui import QTextDocument
            
            # Create printer
            printer = QPrinter(QPrinter.PrinterMode.HighResolution)
            
            # Show print dialog
            print_dialog = QPrintDialog(printer, self)
            if print_dialog.exec() != QDialog.DialogCode.Accepted:
                return
            
            # Create HTML content
            html = "<html><head><style>"
            html += "table { width: 100%; border-collapse: collapse; direction: rtl; }"
            html += "th, td { border: 1px solid #000; padding: 8px; text-align: right; }"
            html += "th { background-color: #f2f2f2; }"
            html += "h1 { text-align: center; color: #2c3e50; }"
            html += "</style></head><body>"
            
            # Add title based on report type
            report_title = self.report_type.currentText()
            html += f"<h1>{report_title}</h1>"
            
            # Add date range
            from_date = self.from_date.date().toString("yyyy-MM-dd")
            to_date = self.to_date.date().toString("yyyy-MM-dd")
            html += f"<p style='text-align: center;'>الفترة من {from_date} إلى {to_date}</p>"
            
            # Create table
            html += "<table><tr>"
            
            # Add headers
            for col in range(self.report_table.columnCount()):
                header_text = self.report_table.horizontalHeaderItem(col).text()
                html += f"<th>{header_text}</th>"
            html += "</tr>"
            
            # Add data rows
            for row in range(self.report_table.rowCount()):
                html += "<tr>"
                for col in range(self.report_table.columnCount()):
                    item = self.report_table.item(row, col)
                    text = item.text() if item else ""
                    html += f"<td>{text}</td>"
                html += "</tr>"
            
            html += "</table></body></html>"
            
            # Print
            document = QTextDocument()
            document.setHtml(html)
            document.print(printer)
            
        except Exception as e:
            print(f"Error printing report: {e}")
            QMessageBox.warning(self, "خطأ", f"تعذر طباعة التقرير: {str(e)}")
    
    def save_settings(self):
        """Save system settings."""
        try:
            # Get settings values
            system_name = self.system_name_input.text()
            company_name = self.company_name_input.text()
            admin_email = self.admin_email_input.text()
            currency = self.currency_input.currentText()
            language = self.language_input.currentText()
            theme = self.theme_input.currentText()
            
            # Validate inputs
            if not system_name or not company_name or not admin_email:
                QMessageBox.warning(self, "تنبيه", "الرجاء ملء جميع الحقول المطلوبة")
                return
                
            # Create settings data
            settings_data = {
                "system_name": system_name,
                "company_name": company_name,
                "admin_email": admin_email,
                "default_currency": currency,
                "language": language,
                "theme": theme
            }
            
            # Save settings to file
            settings_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "config")
            os.makedirs(settings_dir, exist_ok=True)
            settings_file = os.path.join(settings_dir, "settings.json")
            
            with open(settings_file, 'w', encoding='utf-8') as f:
                json.dump(settings_data, f, ensure_ascii=False, indent=4)
                
            QMessageBox.information(self, "نجاح", "تم حفظ الإعدادات بنجاح")
            
        except Exception as e:
            print(f"Error saving settings: {e}")
            QMessageBox.warning(self, "خطأ", f"تعذر حفظ الإعدادات: {str(e)}")
    
    def change_password(self):
        """Change the user's password."""
        old_password = self.old_password_input.text()
        new_password = self.new_password_input.text()
        confirm_password = self.confirm_password_input.text()
        
        if not old_password or not new_password or not confirm_password:
            QMessageBox.warning(self, "تنبيه", "الرجاء ملء جميع حقول كلمة المرور")
            return
        
        if new_password != confirm_password:
            QMessageBox.warning(self, "تنبيه", "كلمة المرور الجديدة وتأكيدها غير متطابقين")
            return
        
        try:
            headers = {"Authorization": f"Bearer {self.token}"} if self.token else {}
            data = {
                "old_password": old_password,
                "new_password": new_password
            }
            response = requests.post(f"{self.api_url}/change-password/", json=data, headers=headers)
            
            if response.status_code == 200:
                QMessageBox.information(self, "نجاح", "تم تغيير كلمة المرور بنجاح")
                self.old_password_input.clear()
                self.new_password_input.clear()
                self.confirm_password_input.clear()
            else:
                error_msg = f"فشل تغيير كلمة المرور: رمز الحالة {response.status_code}"
                try:
                    error_data = response.json()
                    if "detail" in error_data:
                        error_msg = error_data["detail"]
                except:
                    pass
                
                QMessageBox.warning(self, "خطأ", error_msg)
        except Exception as e:
            print(f"Error changing password: {e}")
            QMessageBox.warning(self, "خطأ", f"تعذر تغيير كلمة المرور: {str(e)}")
    
    def create_backup(self):
        """Create a backup of the database."""
        try:
            # Get backup directory from user
            backup_dir = QFileDialog.getExistingDirectory(
                self, "اختر مجلد النسخ الاحتياطي", ""
            )
            
            if not backup_dir:
                return  # User canceled
                
            # Create backup filename with timestamp
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_file = os.path.join(backup_dir, f"system_backup_{timestamp}.zip")
            
            # Request backup from server
            headers = {"Authorization": f"Bearer {self.token}"} if self.token else {}
            response = requests.get(f"{self.api_url}/backup/", headers=headers, stream=True)
            
            if response.status_code == 200:
                # Save backup file
                with open(backup_file, 'wb') as f:
                    for chunk in response.iter_content(chunk_size=8192):
                        f.write(chunk)
                        
                QMessageBox.information(self, "نجاح", f"تم إنشاء النسخة الاحتياطية بنجاح في:\n{backup_file}")
            else:
                error_msg = f"فشل إنشاء النسخة الاحتياطية: رمز الحالة {response.status_code}"
                try:
                    error_data = response.json()
                    if "detail" in error_data:
                        error_msg = error_data["detail"]
                except:
                    pass
                
                QMessageBox.warning(self, "خطأ", error_msg)
                
        except Exception as e:
            print(f"Error creating backup: {e}")
            QMessageBox.warning(self, "خطأ", f"تعذر إنشاء النسخة الاحتياطية: {str(e)}")
    
    def restore_backup(self):
        """Restore from a backup."""
        try:
            # Get backup file from user
            backup_file, _ = QFileDialog.getOpenFileName(
                self, "اختر ملف النسخة الاحتياطية", "", "ملفات ZIP (*.zip)"
            )
            
            if not backup_file:
                return  # User canceled
                
            # Confirm restore
            confirm = QMessageBox.warning(
                self,
                "تأكيد الاستعادة",
                "سيؤدي استعادة النسخة الاحتياطية إلى استبدال جميع البيانات الحالية. هل أنت متأكد من المتابعة؟",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if confirm != QMessageBox.StandardButton.Yes:
                return
                
            # Upload backup file to server
            headers = {"Authorization": f"Bearer {self.token}"} if self.token else {}
            
            with open(backup_file, 'rb') as f:
                files = {'backup_file': (os.path.basename(backup_file), f, 'application/zip')}
                response = requests.post(f"{self.api_url}/restore/", headers=headers, files=files)
            
            if response.status_code == 200:
                QMessageBox.information(
                    self, 
                    "نجاح", 
                    "تم استعادة النسخة الاحتياطية بنجاح. سيتم إعادة تشغيل النظام."
                )
                # In a real application, you would restart the application here
            else:
                error_msg = f"فشل استعادة النسخة الاحتياطية: رمز الحالة {response.status_code}"
                try:
                    error_data = response.json()
                    if "detail" in error_data:
                        error_msg = error_data["detail"]
                except:
                    pass
                
                QMessageBox.warning(self, "خطأ", error_msg)
                
        except Exception as e:
            print(f"Error restoring backup: {e}")
            QMessageBox.warning(self, "خطأ", f"تعذر استعادة النسخة الاحتياطية: {str(e)}")
