from PyQt6.QtWidgets import (
    QGridLayout, QDialog, QLabel, QComboBox, QDateEdit,
    QTableWidget, QTableWidgetItem, QFileDialog, QMessageBox, 
    QVBoxLayout, QHBoxLayout, QHeaderView, QWidget
)
from PyQt6.QtGui import QFont, QColor
from PyQt6.QtCore import Qt, QDate, QThread, pyqtSignal
from ui.custom_widgets import ModernGroupBox, ModernButton
from utils.helpers import get_status_arabic, get_status_color
import requests
from api.client import APIClient

class ReportWorker(QThread):
    result_ready = pyqtSignal(dict, str)
    error_occurred = pyqtSignal(str)
    progress_updated = pyqtSignal(str)
    def __init__(self, url, headers, params, report_type):
        super().__init__()
        self.url = url
        self.headers = headers
        self.params = params
        self.report_type = report_type
        self._is_cancelled = False
    def run(self):
        try:
            self.progress_updated.emit("جاري تحميل التقرير...")
            response = requests.get(self.url, headers=self.headers, params=self.params, timeout=30)
            if self._is_cancelled:
                self.progress_updated.emit("تم إلغاء التحميل.")
                return
            if response.status_code == 200:
                self.result_ready.emit(response.json(), self.report_type)
            else:
                try:
                    error_data = response.json()
                    msg = error_data.get("detail", f"HTTP {response.status_code}")
                except:
                    msg = f"HTTP {response.status_code}"
                self.error_occurred.emit(msg)
        except Exception as e:
            self.error_occurred.emit(str(e))
    def cancel(self):
        self._is_cancelled = True

class ReportHandlerMixin:
    """Mixin class handling report generation and export functionality"""
    
    def __init__(self):
        """Initialize the mixin with required widgets"""
        self.reports_tab = QWidget()
        self.report_table = QTableWidget()
        self.report_type = QComboBox()
        self.from_date = QDateEdit()
        self.to_date = QDateEdit()
        self.report_branch_filter = QComboBox()
        self.parent = None  # Will be set when the mixin is used
        self.token = None
        self.api_url = None
        self.api_client = None
        self.branch_id_to_name = {}
        self._report_worker = None
        self._loading_label = None
    
    def set_parent(self, parent):
        """Set the parent widget for showing messages"""
        self.parent = parent
    
    def set_api_client(self, token, api_url):
        """Set the API client with token and URL"""
        self.token = token
        self.api_url = api_url
        self.api_client = APIClient(token)
    
    def show_message(self, message, duration=3000):
        """Show a message using QMessageBox"""
        if self.parent:
            QMessageBox.information(self.parent, "معلومات", message)
    
    def show_error(self, message):
        """Show an error message using QMessageBox"""
        if self.parent:
            QMessageBox.warning(self.parent, "خطأ", message)
    
    def show_loading(self, message="جاري التحميل..."):
        if self.parent and not self._loading_label:
            self._loading_label = QLabel(message, self.parent)
            self._loading_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self._loading_label.setStyleSheet("background: #fffbe6; color: #e67e22; font-size: 16px; padding: 20px; border-radius: 10px;")
            self._loading_label.setGeometry(0, 0, self.parent.width(), 60)
            self._loading_label.show()
            self.parent.repaint()
    def hide_loading(self):
        if self._loading_label:
            self._loading_label.hide()
            self._loading_label.deleteLater()
            self._loading_label = None
    
    def setup_reports_tab(self):
        """Set up the reports tab."""
        layout = QVBoxLayout()
        
        # Title
        title = QLabel("التقارير")
        title.setFont(QFont("Arial", 18, QFont.Weight.Bold))
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setStyleSheet("color: #2c3e50; margin-bottom: 20px;")
        layout.addWidget(title)
        
        # Report options
        options_group = ModernGroupBox("خيارات التقرير", "#3498db")
        options_layout = QGridLayout()
        
        # Report type
        report_type_label = QLabel("نوع التقرير:")
        options_layout.addWidget(report_type_label, 0, 0)
        
        self.report_type.addItems(["تقرير التحويلات", "تقرير الفروع", "تقرير الموظفين"])
        options_layout.addWidget(self.report_type, 0, 1)
        
        # Date range
        date_range_label = QLabel("نطاق التاريخ:")
        options_layout.addWidget(date_range_label, 1, 0)
        
        date_range_layout = QHBoxLayout()
        
        from_date_label = QLabel("من:")
        date_range_layout.addWidget(from_date_label)
        
        self.from_date.setCalendarPopup(True)
        self.from_date.setDate(QDate.currentDate().addDays(-30))  # Last 30 days
        date_range_layout.addWidget(self.from_date)
        
        to_date_label = QLabel("إلى:")
        date_range_layout.addWidget(to_date_label)
        
        self.to_date.setCalendarPopup(True)
        self.to_date.setDate(QDate.currentDate())
        date_range_layout.addWidget(self.to_date)
        
        options_layout.addLayout(date_range_layout, 1, 1)
        
        # Branch filter
        branch_filter_label = QLabel("الفرع:")
        options_layout.addWidget(branch_filter_label, 2, 0)
        
        self.report_branch_filter = QComboBox()
        options_layout.addWidget(self.report_branch_filter, 2, 1)
        
        # Generate button
        generate_button = ModernButton("إنشاء التقرير", color="#2ecc71")
        generate_button.clicked.connect(self.generate_report)
        options_layout.addWidget(generate_button, 3, 0, 1, 2)
        
        options_group.setLayout(options_layout)
        layout.addWidget(options_group)
        
        # Report preview
        preview_group = ModernGroupBox("معاينة التقرير", "#e74c3c")
        preview_layout = QVBoxLayout()
        
        self.report_table = QTableWidget()
        preview_layout.addWidget(self.report_table)
        
        # Export buttons
        export_layout = QHBoxLayout()
        
        export_pdf_button = ModernButton("تصدير PDF", color="#3498db")
        export_pdf_button.clicked.connect(self.export_pdf)
        export_layout.addWidget(export_pdf_button)
        
        export_excel_button = ModernButton("تصدير Excel", color="#f39c12")
        export_excel_button.clicked.connect(self.export_excel)
        export_layout.addWidget(export_excel_button)
        
        export_print_button = ModernButton("طباعة", color="#9b59b6")
        export_print_button.clicked.connect(self.print_report)
        export_layout.addWidget(export_print_button)
        
        preview_layout.addLayout(export_layout)
        preview_group.setLayout(preview_layout)
        layout.addWidget(preview_group)
        
        self.reports_tab.setLayout(layout)

    def generate_report(self):
        if self._report_worker and self._report_worker.isRunning():
            self._report_worker.cancel()
            self._report_worker.wait()
        report_type_map = {
            "تقرير التحويلات": "transactions",
            "تقرير الفروع": "branch",
            "تقرير الموظفين": "employees"
        }
        report_type = report_type_map.get(self.report_type.currentText(), "transactions")
        start_date = self.from_date.date().toString("yyyy-MM-dd")
        end_date = self.to_date.date().toString("yyyy-MM-dd")
        branch_id = self.report_branch_filter.currentData()
        if branch_id == self.api_url:
            branch_id = None
        headers = {"Authorization": f"Bearer {self.token}"} if self.token else {}
        params = {}
        url = None
        if report_type == "transactions":
            params["start_date"] = start_date
            params["end_date"] = end_date
            if branch_id:
                params["branch_id"] = branch_id
            url = f"{self.api_url}/transactions/"
        elif report_type == "branch":
            params["start_date"] = start_date
            params["end_date"] = end_date
            if branch_id:
                params["branch_id"] = branch_id
            url = f"{self.api_url}/reports/branch/"
        elif report_type == "employees":
            if branch_id:
                params["branch_id"] = branch_id
            url = f"{self.api_url}/users/"
        else:
            self.show_error("نوع التقرير غير مدعوم")
            return
        self.show_loading("جاري تحميل التقرير...")
        self.report_table.setRowCount(0)
        self._report_worker = ReportWorker(url, headers, params, report_type)
        self._report_worker.result_ready.connect(self._on_report_data_ready)
        self._report_worker.error_occurred.connect(self._on_report_error)
        self._report_worker.progress_updated.connect(self._on_report_progress)
        self._report_worker.finished.connect(self.hide_loading)
        self._report_worker.start()

    def _on_report_data_ready(self, data, report_type):
        if report_type == "transactions":
            items = data.get("transactions", [])
        elif report_type == "branch":
            items = []
            branch_report = data.get("branch_report", {})
            for branch_id, stats in branch_report.items():
                branch = self.branch_id_to_name.get(branch_id, f"الفرع {branch_id}")
                items.append({
                    "branch_id": branch_id,
                    "name": branch,
                    "total_syp": stats.get("total_syp", 0),
                    "total_usd": stats.get("total_usd", 0),
                    "count": stats.get("count", 0)
                })
        elif report_type == "employees":
            items = data.get("users", [])
        elif report_type == "daily":
            items = []
            daily_report = data.get("daily_report", {})
            for date_str, stats in daily_report.items():
                items.append({
                    "date": date_str,
                    "total_syp": stats.get("total_syp", 0),
                    "total_usd": stats.get("total_usd", 0),
                    "count": stats.get("count", 0)
                })
        else:
            items = []
        
        # Set up table columns based on report type
        if report_type == "transactions":
            self.report_table.setColumnCount(10)
            self.report_table.setHorizontalHeaderLabels([
                "النوع", "رقم التحويل", "المرسل", "المستلم", "المبلغ", 
                "التاريخ", "الحالة", "الفرع المرسل", "الفرع المستلم", "اسم الموظف"
            ])
        elif report_type == "branch":
            self.report_table.setColumnCount(5)
            self.report_table.setHorizontalHeaderLabels([
                "رقم الفرع", "اسم الفرع", "إجمالي الليرة", "إجمالي الدولار", "عدد التحويلات"
            ])
        elif report_type == "employees":
            self.report_table.setColumnCount(5)
            self.report_table.setHorizontalHeaderLabels([
                "اسم المستخدم", "الدور", "الفرع", "تاريخ الإنشاء", "الحالة"
            ])
        elif report_type == "daily":
            self.report_table.setColumnCount(4)
            self.report_table.setHorizontalHeaderLabels([
                "التاريخ", "إجمالي الليرة", "إجمالي الدولار", "عدد التحويلات"
            ])
        
        self.report_table.setRowCount(len(items))
        
        # Load branch names if not cached
        if not hasattr(self, 'branch_id_to_name') or not self.branch_id_to_name:
            self.branch_id_to_name = {}
            branches_response = self.api_client.get_branches()
            if branches_response.status_code == 200:
                branches = branches_response.json().get("branches", [])
                self.branch_id_to_name = {b["id"]: b["name"] for b in branches}
        
        for i, item in enumerate(items):
            if report_type == "transactions":
                transaction_type = self.determine_transaction_type(item)
                type_item = QTableWidgetItem(transaction_type)
                if transaction_type == "داخلي":
                    type_item.setForeground(QColor(0, 128, 0))
                elif transaction_type == "صادر":
                    type_item.setForeground(QColor(255, 0, 0))
                elif transaction_type == "وارد":
                    type_item.setForeground(QColor(0, 0, 255))
                self.report_table.setItem(i, 0, type_item)
                trans_id = str(item.get("id", ""))
                id_item = QTableWidgetItem(trans_id[:8] + "..." if len(trans_id) > 8 else trans_id)
                id_item.setToolTip(trans_id)
                self.report_table.setItem(i, 1, id_item)
                self.report_table.setItem(i, 2, QTableWidgetItem(item.get("sender", "")))
                self.report_table.setItem(i, 3, QTableWidgetItem(item.get("receiver", "")))
                amount = item.get("amount", 0)
                currency = item.get("currency", "ليرة سورية")
                amount_item = QTableWidgetItem(f"{amount:,.2f} {currency}")
                amount_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                self.report_table.setItem(i, 4, amount_item)
                date_str = item.get("date", "")
                self.report_table.setItem(i, 5, QTableWidgetItem(date_str))
                status = item.get("status", "").lower()
                status_ar = get_status_arabic(status)
                status_item = QTableWidgetItem(status_ar)
                status_item.setBackground(get_status_color(status))
                status_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.report_table.setItem(i, 6, status_item)
                branch_id = item.get("branch_id")
                sending_branch = self.branch_id_to_name.get(branch_id, f"الفرع {branch_id}" if branch_id else "غير معروف")
                self.report_table.setItem(i, 7, QTableWidgetItem(sending_branch))
                dest_branch_id = item.get("destination_branch_id")
                receiving_branch = self.branch_id_to_name.get(dest_branch_id, f"الفرع {dest_branch_id}" if dest_branch_id else "غير معروف")
                self.report_table.setItem(i, 8, QTableWidgetItem(receiving_branch))
                self.report_table.setItem(i, 9, QTableWidgetItem(item.get("employee_name", "")))
                self.report_table.item(i, 0).setData(Qt.ItemDataRole.UserRole, item)
            elif report_type == "branch":
                self.report_table.setItem(i, 0, QTableWidgetItem(str(item.get("branch_id", ""))))
                self.report_table.setItem(i, 1, QTableWidgetItem(item.get("name", "")))
                self.report_table.setItem(i, 2, QTableWidgetItem(f"{item.get('total_syp', 0):,.2f}"))
                self.report_table.setItem(i, 3, QTableWidgetItem(f"{item.get('total_usd', 0):,.2f}"))
                self.report_table.setItem(i, 4, QTableWidgetItem(str(item.get("count", 0))))
            elif report_type == "employees":
                self.report_table.setItem(i, 0, QTableWidgetItem(item.get("username", "")))
                self.report_table.setItem(i, 1, QTableWidgetItem(item.get("role", "")))
                branch_name = self.branch_id_to_name.get(item.get("branch_id"), "غير معروف")
                self.report_table.setItem(i, 2, QTableWidgetItem(branch_name))
                self.report_table.setItem(i, 3, QTableWidgetItem(item.get("created_at", "")))
                status = "نشط" if item.get("role") != "deleted" else "محذوف"
                self.report_table.setItem(i, 4, QTableWidgetItem(status))
            elif report_type == "daily":
                self.report_table.setItem(i, 0, QTableWidgetItem(item.get("date", "")))
                self.report_table.setItem(i, 1, QTableWidgetItem(f"{item.get('total_syp', 0):,.2f}"))
                self.report_table.setItem(i, 2, QTableWidgetItem(f"{item.get('total_usd', 0):,.2f}"))
                self.report_table.setItem(i, 3, QTableWidgetItem(str(item.get("count", 0))))
        
        self.report_table.horizontalHeader().setStretchLastSection(True)
        self.report_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.report_table.setSortingEnabled(True)
        self.show_message("تم إنشاء التقرير بنجاح")
        QMessageBox.information(self.parent, "نجاح", "تم إنشاء التقرير بنجاح")

    def _on_report_error(self, msg):
        self.show_error(msg)

    def _on_report_progress(self, msg):
        if self._loading_label:
            self._loading_label.setText(msg)

    def export_pdf(self):
        """Export the current report as PDF with loading indicator."""
        if self.report_table.rowCount() == 0:
            QMessageBox.warning(self, "تنبيه", "لا توجد بيانات للتصدير. الرجاء إنشاء تقرير أولاً.")
            return
        try:
            self.show_loading("جاري تصدير التقرير إلى PDF...")
            self._set_export_buttons_enabled(False)
            from PyQt6.QtPrintSupport import QPrinter
            from PyQt6.QtGui import QTextDocument
            file_path, _ = QFileDialog.getSaveFileName(
                self, "حفظ التقرير كـ PDF", "", "ملفات PDF (*.pdf)"
            )
            if not file_path:
                self.hide_loading()
                self._set_export_buttons_enabled(True)
                return
            if not file_path.lower().endswith('.pdf'):
                file_path += '.pdf'
            printer = QPrinter(QPrinter.PrinterMode.HighResolution)
            printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
            printer.setOutputFileName(file_path)
            html = "<html><head><style>"
            html += "table { width: 100%; border-collapse: collapse; direction: rtl; }"
            html += "th, td { border: 1px solid #000; padding: 8px; text-align: right; }"
            html += "th { background-color: #f2f2f2; }"
            html += "h1 { text-align: center; color: #2c3e50; }"
            html += ".status-completed { background-color: #d5f5e3; }"
            html += ".status-processing { background-color: #d6eaf8; }"
            html += ".status-cancelled { background-color: #f5b7b1; }"
            html += ".status-rejected { background-color: #f1948a; }"
            html += ".status-on_hold { background-color: #fdebd0; }"
            html += "</style></head><body>"
            report_title = self.report_type.currentText()
            html += f"<h1>{report_title}</h1>"
            from_date = self.from_date.date().toString("yyyy-MM-dd")
            to_date = self.to_date.date().toString("yyyy-MM-dd")
            html += f"<p style='text-align: center;'>الفترة من {from_date} إلى {to_date}</p>"
            html += "<table><tr>"
            for col in range(self.report_table.columnCount()):
                header_text = self.report_table.horizontalHeaderItem(col).text()
                html += f"<th>{header_text}</th>"
            html += "</tr>"
            for row in range(self.report_table.rowCount()):
                html += "<tr>"
                status_item = None
                status_class = ""
                if self.report_type.currentText() == "تقرير التحويلات":
                    status_item = self.report_table.item(row, 6)
                    if status_item:
                        status_text = status_item.text().lower()
                        if "مكتمل" in status_text:
                            status_class = "status-completed"
                        elif "قيد المعالجة" in status_text:
                            status_class = "status-processing"
                        elif "ملغي" in status_text:
                            status_class = "status-cancelled"
                        elif "مرفوض" in status_text:
                            status_class = "status-rejected"
                        elif "معلق" in status_text:
                            status_class = "status-on_hold"
                for col in range(self.report_table.columnCount()):
                    item = self.report_table.item(row, col)
                    text = item.text() if item else ""
                    if self.report_type.currentText() == "تقرير التحويلات" and status_class:
                        html += f"<td class='{status_class}'>{text}</td>"
                    else:
                        html += f"<td>{text}</td>"
                html += "</tr>"
            html += "</table></body></html>"
            document = QTextDocument()
            document.setHtml(html)
            document.print(printer)
            self.hide_loading()
            self._set_export_buttons_enabled(True)
            QMessageBox.information(self, "نجاح", f"تم تصدير التقرير بنجاح إلى:\n{file_path}")
        except Exception as e:
            self.hide_loading()
            self._set_export_buttons_enabled(True)
            print(f"Error exporting to PDF: {e}")
            QMessageBox.warning(self, "خطأ", f"تعذر تصدير التقرير: {str(e)}")

    def export_excel(self):
        """Export the current report as Excel with loading indicator."""
        if self.report_table.rowCount() == 0:
            QMessageBox.warning(self, "تنبيه", "لا توجد بيانات للتصدير. الرجاء إنشاء تقرير أولاً.")
            return
        try:
            self.show_loading("جاري تصدير التقرير إلى Excel...")
            self._set_export_buttons_enabled(False)
            file_path, _ = QFileDialog.getSaveFileName(
                self, "حفظ التقرير كـ Excel", "", "ملفات Excel (*.xlsx)"
            )
            if not file_path:
                self.hide_loading()
                self._set_export_buttons_enabled(True)
                return
            if not file_path.lower().endswith('.xlsx'):
                file_path += '.xlsx'
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            wb = openpyxl.Workbook()
            ws = wb.active
            report_title = self.report_type.currentText()
            ws.title = report_title
            ws.merge_cells('A1:G1')
            title_cell = ws['A1']
            title_cell.value = report_title
            title_cell.font = Font(size=16, bold=True)
            title_cell.alignment = Alignment(horizontal='center')
            from_date = self.from_date.date().toString("yyyy-MM-dd")
            to_date = self.to_date.date().toString("yyyy-MM-dd")
            ws.merge_cells('A2:G2')
            date_cell = ws['A2']
            date_cell.value = f"الفترة من {from_date} إلى {to_date}"
            date_cell.alignment = Alignment(horizontal='center')
            header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            header_font = Font(bold=True)
            for col in range(self.report_table.columnCount()):
                header_text = self.report_table.horizontalHeaderItem(col).text()
                cell = ws.cell(row=4, column=col+1, value=header_text)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='right')
            for row in range(self.report_table.rowCount()):
                for col in range(self.report_table.columnCount()):
                    item = self.report_table.item(row, col)
                    text = item.text() if item else ""
                    cell = ws.cell(row=row+5, column=col+1, value=text)
                    cell.alignment = Alignment(horizontal='right')
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            wb.save(file_path)
            self.hide_loading()
            self._set_export_buttons_enabled(True)
            QMessageBox.information(self, "نجاح", f"تم تصدير التقرير بنجاح إلى:\n{file_path}")
        except ImportError:
            self.hide_loading()
            self._set_export_buttons_enabled(True)
            QMessageBox.warning(self, "خطأ", "مكتبة openpyxl غير متوفرة. الرجاء تثبيتها باستخدام pip install openpyxl")
        except Exception as e:
            self.hide_loading()
            self._set_export_buttons_enabled(True)
            print(f"Error exporting to Excel: {e}")
            QMessageBox.warning(self, "خطأ", f"تعذر تصدير التقرير: {str(e)}")

    def _set_export_buttons_enabled(self, enabled: bool):
        """Enable or disable export buttons during export operations."""
        if hasattr(self, 'reports_tab'):
            for btn_text in ["تصدير PDF", "تصدير Excel", "طباعة"]:
                btns = self.reports_tab.findChildren(ModernButton, btn_text)
                for btn in btns:
                    btn.setEnabled(enabled)

    def print_report(self):
        """Print the current report with loading indicator."""
        if self.report_table.rowCount() == 0:
            QMessageBox.warning(self, "تنبيه", "لا توجد بيانات للطباعة. الرجاء إنشاء تقرير أولاً.")
            return
        try:
            self.show_loading("جاري تجهيز التقرير للطباعة...")
            self._set_export_buttons_enabled(False)
            from PyQt6.QtPrintSupport import QPrinter, QPrintDialog
            from PyQt6.QtGui import QTextDocument
            printer = QPrinter(QPrinter.PrinterMode.HighResolution)
            print_dialog = QPrintDialog(printer, self)
            if print_dialog.exec() != QDialog.DialogCode.Accepted:
                self.hide_loading()
                self._set_export_buttons_enabled(True)
                return
            html = "<html><head><style>"
            html += "table { width: 100%; border-collapse: collapse; direction: rtl; }"
            html += "th, td { border: 1px solid #000; padding: 8px; text-align: right; }"
            html += "th { background-color: #f2f2f2; }"
            html += "h1 { text-align: center; color: #2c3e50; }"
            html += "</style></head><body>"
            report_title = self.report_type.currentText()
            html += f"<h1>{report_title}</h1>"
            from_date = self.from_date.date().toString("yyyy-MM-dd")
            to_date = self.to_date.date().toString("yyyy-MM-dd")
            html += f"<p style='text-align: center;'>الفترة من {from_date} إلى {to_date}</p>"
            html += "<table><tr>"
            for col in range(self.report_table.columnCount()):
                header_text = self.report_table.horizontalHeaderItem(col).text()
                html += f"<th>{header_text}</th>"
            html += "</tr>"
            for row in range(self.report_table.rowCount()):
                html += "<tr>"
                for col in range(self.report_table.columnCount()):
                    item = self.report_table.item(row, col)
                    text = item.text() if item else ""
                    html += f"<td>{text}</td>"
                html += "</tr>"
            html += "</table></body></html>"
            document = QTextDocument()
            document.setHtml(html)
            document.print(printer)
            self.hide_loading()
            self._set_export_buttons_enabled(True)
        except Exception as e:
            self.hide_loading()
            self._set_export_buttons_enabled(True)
            print(f"Error printing report: {e}")
            QMessageBox.warning(self, "خطأ", f"تعذر طباعة التقرير: {str(e)}")

    def determine_transaction_type(self, transaction):
        """Determine transaction type based on branch information."""
        branch_id = transaction.get("branch_id")
        dest_branch_id = transaction.get("destination_branch_id")
        
        if branch_id and dest_branch_id:
            return "داخلي"  # Internal transfer between branches
        elif branch_id and not dest_branch_id:
            return "صادر"   # Outgoing transfer from branch
        elif not branch_id and dest_branch_id:
            return "وارد"   # Incoming transfer to branch
        else:
            return "غير معروف"  # Unknown type